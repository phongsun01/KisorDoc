import asyncio
import errno
import sys
import threading
import time
import traceback
from datetime import datetime
from pathlib import Path
import pandas as pd
import socket
import re
import math
import webbrowser
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

sys.stdout.reconfigure(encoding="utf-8")

import gradio as gr

from config import load_config, AppConfig
from dataset import DataSet
from file_utils import clear_output_folder, copy_templates_to_output, rename_output, open_output_folder, cleanup_old_logs, open_logs_folder
from merger import mail_merge_safe
from table_copier import copy_tables_for_file
import shutil
from docxtpl import DocxTemplate

config: AppConfig | None = None
ds: DataSet | None = None


def _str(val, default=""):
    if val is None:
        return default
    if isinstance(val, float):
        if math.isnan(val):
            return default
    return str(val).strip()


def clean_config_key(key: str) -> str:
    clean = key.strip("<>{}| ")
    
    # Maps dot date suffixes to _Date suffix
    for suffix in (".Date.Long", ".Date.long", ".date_long"):
        if clean.endswith(suffix):
            return clean[:-len(suffix)] + "_Date"
            
    if clean.endswith(".Date") or clean.endswith(".date"):
        return clean[:-5] + "_Date"
        
    if clean.endswith(".Day") or clean.endswith(".day"):
        return clean[:-4] + "_Date"
        
    if clean.endswith(".Month") or clean.endswith(".month"):
        return clean[:-6] + "_Date"
        
    if clean.endswith(".Year") or clean.endswith(".year"):
        return clean[:-5] + "_Date"
        
    # Other standard modifiers
    for suffix in (".Upper", ".upper", ".Number", ".number"):
        if clean.endswith(suffix):
            clean = clean[:-len(suffix)]
            break
            
    if "|" in clean:
        clean = clean.split("|")[0].strip()
        
    return clean


def init():
    global config, ds
    config = load_config()
    ds = DataSet(config)
    cleanup_old_logs(config)


def get_options() -> list[str]:
    rows = ds.query("SELECT Key, Value FROM Options ORDER BY Key")
    return [f"{_str(r['Key'])}: {_str(r['Value'])}" for r in rows]


def get_option_config(option_key: str) -> dict:
    if not option_key:
        return {}
    opt_code = option_key.split(":")[0].strip() if ":" in option_key else option_key.strip()
    try:
        rows = ds.query("SELECT * FROM Options")
    except Exception:
        rows = []
    for r in rows:
        if _str(r.get("Key")) == opt_code:
            return {
                "sheet": _str(r.get("Sheet"), "GoiThau"),
                "show": _str(r.get("Show"), "{TT}. {Số hiệu gói thầu} - {Tên gói thầu}"),
                "key_id": _str(r.get("KeyId"), "GoiThau_ID"),
                "config_range": _str(r.get("Config"), ""),
            }
    return {
        "sheet": "GoiThau",
        "show": "{TT}. {Số hiệu gói thầu} - {Tên gói thầu}",
        "key_id": "GoiThau_ID",
        "config_range": "",
    }


def _parse_row_range(s: str) -> tuple[int, int] | None:
    if not s or not s.strip():
        return None
    import re
    m = re.match(r"^(\d+)-(\d+)$", s.strip())
    if m:
        try:
            start = int(m.group(1))
            end = int(m.group(2))
            if start <= end:
                return start, end
        except ValueError:
            pass
    return None


def get_config_for_option(option_key: str) -> list[dict]:
    opt_config = get_option_config(option_key)
    cfg_range = opt_config.get("config_range", "")
    parsed = _parse_row_range(cfg_range)
    if parsed:
        start, end = parsed
        return ds.query_rows("Config", start, end)
    try:
        return ds.query("SELECT * FROM Config")
    except Exception:
        return []


def get_all_option_templates(option_key: str) -> list[str]:
    if not option_key:
        return []
    opt = option_key.split(":")[0].strip() if ":" in option_key else option_key.strip()
    try:
        ws_rows = ds.query("SELECT * FROM Workflow")
    except Exception:
        ws_rows = []
    return [str(r.get("Name", "")) for r in ws_rows if _str(r.get("Option")) == opt]


def safe_format(pattern: str, row: dict) -> str:
    if not pattern:
        return ""
    import re
    res = pattern
    placeholders = re.findall(r"\{(.*?)\}", pattern)
    for p in placeholders:
        val = _str(row.get(p, ""))
        res = res.replace(f"{{{p}}}", val)
    return res


def check_condition(condition_str: str, raw_row: dict, config_mappings: list[dict]) -> bool:
    if not condition_str or condition_str.strip() == "" or condition_str.upper() == "ALL":
        return True
    
    import re
    placeholders = re.findall(r"\{(.*?)\}", condition_str)
    
    key_to_col = {}
    for mapping in config_mappings:
        k = clean_config_key(_str(mapping.get("Key")))
        c = _str(mapping.get("Value"))
        if k and c:
            key_to_col[k] = c

    eval_context = {}
    patched_condition = condition_str
    
    for idx, p in enumerate(placeholders):
        p_clean = p.strip()
        safe_var = f"var_{idx}"
        
        col_name = key_to_col.get(p_clean, p_clean)
        val = raw_row.get(col_name)
        
        parsed_val = None
        if val is not None:
            try:
                is_na = pd.isna(val)
            except (TypeError, ValueError):
                is_na = False

            if not is_na:
                parsed_val = _parse_price(val)
                if parsed_val is None:
                    s_val = str(val).strip()
                    if s_val:
                        parsed_val = s_val
        
        eval_context[safe_var] = parsed_val
        patched_condition = patched_condition.replace(f"{{{p}}}", safe_var)

    try:
        return bool(eval(patched_condition, {}, eval_context))
    except Exception as e:
        print(f"⚠️  Lỗi cú pháp điều kiện lọc '{condition_str}' (biểu thức sau chuyển đổi: '{patched_condition}'): {e}")
        return False


def parse_join_expression(expr: str) -> str:
    """
    Cú pháp rút gọn cho cột Sheet trong Options:
      Table1 <* Table2 @ key           → LEFT JOIN, cùng tên cột
      Table1 <* Table2 @ key1 = key2   → LEFT JOIN, khác tên cột
      Table1 *> Table2 @ key           → RIGHT JOIN
      Table1 * Table2 @ key            → INNER JOIN
      Table1 <*> Table2 @ key          → FULL OUTER JOIN
      SELECT ...                        → truyền thẳng cho DuckDB
    """
    s = expr.strip()
    if s.lower().startswith("select"):
        return s
    if "@" not in s:
        return f'SELECT * FROM "{s}"'

    join_part, key_raw = s.split("@", 1)
    join_part = join_part.strip()
    key_raw   = key_raw.strip()

    _OP_MAP = [
        (" <*>", "FULL OUTER JOIN"),
        (" <*",  "LEFT JOIN"),
        (" *>",  "RIGHT JOIN"),
        (" *",   "INNER JOIN"),
    ]
    join_type = None
    t1 = t2 = ""
    for sym, jt in _OP_MAP:
        if sym in join_part:
            join_type = jt
            left, right = join_part.split(sym.strip(), 1)
            t1 = left.strip()
            t2 = right.strip()
            break

    if not join_type:
        return f'SELECT * FROM "{s}"'

    if "=" in key_raw:
        k1, k2 = [k.strip() for k in key_raw.split("=", 1)]
    else:
        k1 = k2 = key_raw

    return (
        f'SELECT * FROM "{t1}" {join_type} "{t2}" '
        f'ON "{t1}"."{k1}" = "{t2}"."{k2}"'
    )


_JOIN_RE = re.compile(r'.+\s+(?:<\*>|<\*|\*>|\*)\s+.+\s*@\s*.+', re.DOTALL)


def resolve_sheet_query(sheet_name: str) -> str:
    """
    Chuyển đổi giá trị cột Sheet trong Options thành SQL:
    - Bắt đầu bằng SELECT → passthrough
    - Khớp pattern join rút gọn → gọi parse_join_expression
    - Còn lại → SELECT * FROM "<sheet_name>"
    """
    s = sheet_name.strip()
    if not s:
        return f'SELECT * FROM "{s}"'
    if s.lower().startswith("select"):
        return s
    if _JOIN_RE.match(s):
        return parse_join_expression(s)
    return f'SELECT * FROM "{s}"'


def get_packages(option_key: str) -> list[str]:
    if not option_key:
        return []
    opt_config = get_option_config(option_key)
    sheet = opt_config.get("sheet", "GoiThau")
    show_format = opt_config.get("show", "")
    
    sql = resolve_sheet_query(sheet)
    try:
        rows = ds.query(f"SELECT * FROM ({sql}) ORDER BY CAST(TT AS INTEGER)")
    except Exception:
        try:
            rows = ds.query(sql)
        except Exception:
            rows = []
        
    packages = []
    for r in rows:
        label = safe_format(show_format, r)
        if label:
            packages.append(label)
    return packages


def get_package_details(option_key: str, package_label: str, sheet_rows: list[dict] | None = None) -> dict:
    if not option_key or not package_label:
        return {}
    opt_config = get_option_config(option_key)
    show_format = opt_config.get("show", "")

    if sheet_rows is not None:
        rows = sheet_rows
    else:
        sheet = opt_config.get("sheet", "GoiThau")
        sql = resolve_sheet_query(sheet)
        try:
            rows = ds.query(sql)
        except Exception:
            rows = []
        
    for r in rows:
        label = safe_format(show_format, r)
        if label == package_label:
            details = {}
            for col_name, val in r.items():
                if val is not None and str(val).strip() != "":
                    details[col_name] = _str(val)
            return details
    return {}


def get_workflow_templates(option_key: str, package_label: str, sheet_rows: list[dict] | None = None) -> list[dict]:
    if not option_key or not package_label or package_label.strip() == "":
        return []
    opt = option_key.split(":")[0].strip() if ":" in option_key else option_key.strip()
    
    try:
        ws_rows = ds.query("SELECT * FROM Workflow")
    except Exception:
        ws_rows = []
    wf_rows = [r for r in ws_rows if _str(r.get("Option")) == opt]
    if not wf_rows:
        return []
        
    opt_config = get_option_config(option_key)
    key_id = opt_config.get("key_id", "GoiThau_ID")
    show_format = opt_config.get("show", "")
    
    if sheet_rows is not None:
        main_rows = sheet_rows
    else:
        sheet = opt_config.get("sheet", "GoiThau")
        sql = resolve_sheet_query(sheet)
        try:
            main_rows = ds.query(sql)
        except Exception:
            main_rows = []
        
    selected_pkg = None
    for r in main_rows:
        label = safe_format(show_format, r)
        if label == package_label:
            selected_pkg = r
            break
            
    if not selected_pkg:
        return []
        
    config_mappings = get_config_for_option(option_key)
        
    filtered = []
    for r in wf_rows:
        condition_str = _str(r.get("Condition", ""))
        if not condition_str:
            price = _parse_price(selected_pkg.get("Giá gói thầu", 0))
            goi_thau_loai = _str(selected_pkg.get("GoiThau_Loai"))
            pmin = _parse_price(r.get("Price", 0))
            pmax = _parse_price(r.get("PriceMax", 0))
            if pmin is not None and pmax is not None and price is not None:
                if not (pmin <= price <= pmax):
                    continue
            rtype = _str(r.get("Type"))
            if rtype != "ALL" and rtype != goi_thau_loai:
                continue
            filtered.append(r)
        else:
            if check_condition(condition_str, selected_pkg, config_mappings):
                filtered.append(r)
    return filtered


def _parse_price(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if math.isnan(val):
            return None
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


class NestedVal(dict):
    def __init__(self, val, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self["_val"] = val

    def __str__(self):
        return str(self.get("_val", ""))

    def __repr__(self):
        return self.__str__()

    def __html__(self):
        return self.__str__()


def make_nested_dict(flat_dict: dict) -> dict:
    nested = {}
    for key, value in flat_dict.items():
        parts = key.split(".")
        d = nested
        for part in parts[:-1]:
            if part not in d:
                d[part] = {}
            elif not isinstance(d[part], dict):
                d[part] = NestedVal(d[part])
            d = d[part]
            
        last_part = parts[-1]
        if last_part in d:
            if isinstance(d[last_part], dict):
                d[last_part]["_val"] = value
            else:
                d[last_part] = value
        else:
            d[last_part] = value
            
    # Post-process: Convert any standard dict that has "_val" key into NestedVal
    def convert_to_nested_val(obj):
        if isinstance(obj, dict):
            for k, v in list(obj.items()):
                obj[k] = convert_to_nested_val(v)
            if "_val" in obj and not isinstance(obj, NestedVal):
                nv = NestedVal(obj["_val"])
                for k, v in obj.items():
                    if k != "_val":
                        nv[k] = v
                return nv
        return obj

    for k, v in list(nested.items()):
        nested[k] = convert_to_nested_val(v)
        
    return nested


def run_preview(option_key: str, package_label: str, 
                selected_templates: list[str]) -> str:
    """
    Simplified dry-run: kiểm tra context + tables mà không tạo file.
    Trả về string hiển thị trong preview_box.
    """
    if not option_key or not package_label or not selected_templates:
        return "⚠️ Chọn đủ quy trình, gói thầu và ít nhất 1 template trước"

    opt = option_key.split(":")[0].strip() if ":" in option_key else option_key

    # --- Build context (copy từ run_batch) ---
    opt_config = get_option_config(option_key)
    sheet = opt_config.get("sheet", "GoiThau")
    key_id = opt_config.get("key_id", "GoiThau_ID")
    show_format = opt_config.get("show", "")

    sql = resolve_sheet_query(sheet)
    goi_thau_rows = ds.query(sql)
    selected_pkg = next((
        r for r in goi_thau_rows
        if safe_format(show_format, r) == package_label
    ), None)
    if not selected_pkg:
        return "❌ Không tìm thấy dòng dữ liệu tương ứng"

    config_rows = get_config_for_option(option_key)
    context_keys: set[str] = set()
    missing_keys: list[str] = []

    for r in config_rows:
        key = _str(r.get("Key"))
        col = _str(r.get("Value"))
        if not key or not col:
            continue
        clean_key = clean_config_key(key)

        raw_value = selected_pkg.get(col)
        try:
            is_na = pd.isna(raw_value)
        except (TypeError, ValueError):
            is_na = False

        context_keys.add(clean_key)
        if is_na or raw_value is None or str(raw_value).strip() == "":
            missing_keys.append(clean_key)

    lines = []

    # Dòng 1: context summary
    if missing_keys:
        lines.append(
            f"✅ Context: {len(context_keys)} key  |  "
            f"⚠️ Thiếu data: {', '.join(missing_keys)}"
        )
    else:
        lines.append(f"✅ Context: {len(context_keys)} key – đầy đủ")

    # --- Kiểm tra Tables ---
    try:
        tables_rows = ds.query("SELECT * FROM Tables")
    except Exception:
        tables_rows = []

    goi_thau_id = _str(selected_pkg.get(key_id))

    # Tìm các dòng Tables liên quan đến template đang chọn
    xlsx_files = sorted(config.data_path.glob("*.xlsx"))
    danh_muc_file = next(
        (f for f in xlsx_files if "DanhMuc" in f.stem or "danh muc" in f.stem.lower()),
        xlsx_files[0] if xlsx_files else None
    )

    wb = None
    if danh_muc_file and danh_muc_file.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(danh_muc_file, read_only=True, data_only=True)
        except Exception:
            wb = None

    table_lines = []
    for t in tables_rows:
        t_id = t.get(key_id) if key_id in t else t.get("GoiThau_ID")
        if _str(t_id) != goi_thau_id:
            continue
        name     = _str(t.get("Name", ""))
        tbl_sheet = _str(t.get("Sheet", ""))
        range_   = _str(t.get("Range", ""))
        hide     = _str(t.get("Hide", ""))

        # Đếm số dòng thực tế trong sheet nếu có file
        row_count = "?"
        if wb and tbl_sheet and tbl_sheet in wb.sheetnames:
            try:
                ws = wb[tbl_sheet]
                # Tính max_row trong range
                if ":" in range_:
                    parts = range_.split(":")
                    end = parts[1]
                    digits = "".join(c for c in end if c.isdigit())
                    row_count = digits if digits else str(ws.max_row)
                else:
                    row_count = str(ws.max_row)
            except Exception:
                pass

        hide_str = f"  |  ẩn: {hide}" if hide else ""
        table_lines.append(
            f"📋 {name} → {tbl_sheet}  {range_}  {row_count} dòng{hide_str}"
        )

    if wb:
        try:
            wb.close()
        except Exception:
            pass

    if table_lines:
        lines.extend(table_lines)
    elif tables_rows:
        lines.append("📋 Tables: không có dòng nào khớp với gói thầu này")
    else:
        lines.append("📋 Tables: (không có sheet Tables)")

    return "\n".join(lines)


def write_with_retry(func, max_retries=3, delay=2.0, yield_fn=None):
    for attempt in range(1, max_retries + 1):
        try:
            return func()
        except PermissionError as e:
            if "being used by another process" in str(e) or e.errno == errno.EACCES:
                if attempt == max_retries:
                    raise
                msg = f"🔒 File bị khóa – thử lại lần {attempt}/{max_retries} sau {delay:.0f}s..."
                if yield_fn:
                    yield_fn(msg)
                time.sleep(delay)
            else:
                raise


class IncrementalRunLogger:
    def __init__(self, config, goi_thau_id, option, mode="run"):
        self.config = config
        self.goi_thau_id = goi_thau_id
        self.option = option
        self.mode = mode
        
        self.log_dir = Path(config.ProjectPath) / "logs"
        self.log_dir.mkdir(exist_ok=True)
        
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        suffix = f"_{mode}" if mode != "run" else ""
        self.filepath = self.log_dir / f"{ts}_{goi_thau_id}_{option}{suffix}.log"
        self.start_time = time.time()
        self.ok_count = 0
        self.error_count = 0
        self.warning_count = 0
        self._fh = open(self.filepath, "w", encoding="utf-8-sig")

    def write_header(self, option_desc, package_desc, total_files):
        time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        mode_str = "Chạy thật" if self.mode == "run" else ("Retry" if self.mode == "retry" else "Dry-run")
        header = f"""=====================================
KisorDoc – Run Log
=====================================
Thời gian     : {time_str}
Option        : {option_desc}
Gói thầu      : {package_desc}
Chế độ        : {mode_str}
Tổng file     : {total_files}
=====================================

"""
        self._fh.write(header)
        self._fh.flush()

    def log_event(self, emoji, name, extra_info=None):
        time_str = datetime.now().strftime("%H:%M:%S")
        line = f"[{time_str}] {emoji} {name}\n"
        if extra_info:
            line += f"           {extra_info}\n"
        self._fh.write(line)
        self._fh.flush()

    def write_footer(self):
        elapsed = time.time() - self.start_time
        footer = f"""
=====================================
Kết quả: {self.ok_count} thành công / {self.error_count} lỗi / {self.warning_count} warning
Thời gian chạy: {elapsed:.1f} giây
=====================================
"""
        self._fh.write(footer)
        self._fh.close()


async def run_batch(option_key: str, package_label: str, selected_templates: list[str],
                          progress: gr.Progress = gr.Progress(), retry_state: dict | None = None):

    if not option_key or not option_key.strip():
        yield "", "⚠️ Vui lòng chọn quy trình", None
        return
    if not package_label or not package_label.strip():
        yield "", "⚠️ Vui lòng chọn gói thầu", None
        return
    if not selected_templates or len(selected_templates) == 0:
        yield "", "⚠️ Vui lòng chọn ít nhất 1 template", None
        return

    opt = option_key.split(":")[0].strip() if ":" in option_key else option_key.strip()

    opt_config = get_option_config(option_key)
    sheet = opt_config.get("sheet", "GoiThau")
    key_id = opt_config.get("key_id", "GoiThau_ID")
    show_format = opt_config.get("show", "")

    sql = resolve_sheet_query(sheet)
    goi_thau_rows = ds.query(sql)
    selected_pkg = None
    for r in goi_thau_rows:
        label = safe_format(show_format, r)
        if label == package_label:
            selected_pkg = r
            break

    if not selected_pkg:
        yield "", "❌ Không tìm thấy dòng dữ liệu tương ứng", None
        return

    goi_thau_id = _str(selected_pkg.get(key_id))
    config_rows = get_config_for_option(option_key)

    try:
        tables_rows = ds.query("SELECT * FROM Tables")
    except Exception:
        tables_rows = []

    context = {}
    for r in config_rows:
        key = _str(r.get("Key"))
        col = _str(r.get("Value"))
        if not key or not col:
            continue
        clean_key = clean_config_key(key)

        raw_value = selected_pkg.get(col, "")

        try:
            is_na = pd.isna(raw_value)
        except (TypeError, ValueError):
            is_na = False
        if is_na:
            raw_value = ""
        elif isinstance(raw_value, datetime):
            raw_value = raw_value.strftime("%d/%m/%Y")
        elif raw_value is None:
            raw_value = ""

        context[clean_key] = str(raw_value)

    nested_context = make_nested_dict(context)
    nested_context["now"] = datetime.now()

    xlsx_files = sorted(config.data_path.glob("*.xlsx"))
    danh_muc_file = next(
        (f for f in xlsx_files if "DanhMuc" in f.stem or "danh muc" in f.stem.lower()),
        xlsx_files[0] if xlsx_files else None
    )

    if progress:
        progress(0, desc="Bắt đầu xử lý...")
    yield "", "Bắt đầu...", None

    mode = "retry" if retry_state else "run"
    logger = IncrementalRunLogger(config, goi_thau_id, opt, mode)

    # F3+F6: Nếu là retry, chỉ xử lý các template bị lỗi; nếu không thì xóa folder và chạy lại
    if retry_state:
        # Retry: chỉ copy và xử lý các template bị lỗi
        failed_names = retry_state.get("failed_templates", [])
        template_filenames, template_names = [], []
        for r in get_workflow_templates(option_key, package_label):
            if r.get("Name", "") in failed_names:
                fname_raw = _str(r.get("File", ""))
                fname = fname_raw if fname_raw.endswith(".docx") else fname_raw + ".docx"
                template_filenames.append(fname)
                template_names.append(r.get("Name", ""))
        # Copy template bị lỗi vào output (không xóa folder) with retry
        copied = []
        template_names_actual = []
        src_dir = config.template_path / opt
        for fname, tname in zip(template_filenames, template_names):
            src = src_dir / fname
            if not src.exists():
                possible = list(src_dir.glob(f"{fname}*"))
                if possible:
                    src = possible[0]
                else:
                    continue
            dst = config.output_path / src.name
            def do_copy(s=src, d=dst):
                shutil.copy2(s, d)
            try:
                write_with_retry(do_copy, max_retries=3, delay=2.0)
                copied.append(dst)
                template_names_actual.append(tname)
            except PermissionError:
                # Skip file locked during copy – will be caught in processing loop
                copied.append(dst)
                template_names_actual.append(tname)
        template_names = template_names_actual
    else:
        # Run mới: xóa folder và copy tất cả
        clear_output_folder(config)
        template_filenames, template_names = [], []
        for r in get_workflow_templates(option_key, package_label):
            if r.get("Name", "") in selected_templates:
                fname_raw = _str(r.get("File", ""))
                fname = fname_raw if fname_raw.endswith(".docx") else fname_raw + ".docx"
                template_filenames.append(fname)
                template_names.append(r.get("Name", ""))
        copied = copy_templates_to_output(config, opt, template_filenames)

    total = len(copied)
    logger.write_header(option_key, package_label, total)

    results = []
    used_names: set[str] = set()
    has_locked = False
    has_other_error = False
    failed_templates: list[str] = []

    # Lấy danh sách table placeholder names từ tables_rows
    table_placeholder_names = {
        _str(t.get("Name", "")).strip("{} ")
        for t in tables_rows
        if _str(t.get(key_id) if key_id in t else t.get("GoiThau_ID")) == goi_thau_id
    }

    for i, (src_path, tpl_name) in enumerate(zip(copied, template_names)):
        if progress:
            progress((i + 1) / total, desc=f"Đang xử lý: {tpl_name}")
        try:
            # Check for missing placeholders
            missing_placeholders = []
            try:
                import jinja2
                from filters import (
                    filter_date, filter_date_long, filter_number, filter_num2text,
                    filter_day, filter_month, filter_year, filter_add_days,
                    filter_add_months, filter_date_diff, filter_quarter,
                    filter_weekday, filter_date_text
                )
                jenv = jinja2.Environment()
                jenv.filters["date"] = filter_date
                jenv.filters["date_long"] = filter_date_long
                jenv.filters["number"] = filter_number
                jenv.filters["num2text"] = filter_num2text
                jenv.filters["day"] = filter_day
                jenv.filters["month"] = filter_month
                jenv.filters["year"] = filter_year
                jenv.filters["add_days"] = filter_add_days
                jenv.filters["add_months"] = filter_add_months
                jenv.filters["date_diff"] = filter_date_diff
                jenv.filters["quarter"] = filter_quarter
                jenv.filters["weekday"] = filter_weekday
                jenv.filters["date_text"] = filter_date_text

                doc = DocxTemplate(str(src_path))
                undeclared = doc.get_undeclared_template_variables(jinja_env=jenv)
                for var in undeclared:
                    var_clean = var.strip()
                    if var_clean not in nested_context and var_clean not in table_placeholder_names:
                        missing_placeholders.append(var_clean)
            except Exception as e:
                print(f"⚠️  Lỗi quét placeholder cho {tpl_name}: {e}")

            def do_merge(s=src_path):
                return mail_merge_safe(s, nested_context, s)

            def on_locked_retry(msg):
                if progress:
                    progress((i + 1) / total, desc=f"{tpl_name}: {msg}")

            ok, err = write_with_retry(do_merge, max_retries=3, delay=2.0, yield_fn=on_locked_retry)
            if not ok:
                raise RuntimeError(err)

            if danh_muc_file and danh_muc_file.exists():
                try:
                    copy_tables_for_file(src_path, config, goi_thau_id, tables_rows, danh_muc_file, key_id)
                except PermissionError as table_err:
                    if "being used by another process" in str(table_err) or getattr(table_err, 'errno', None) == errno.EACCES:
                        raise PermissionError(table_err) from table_err
                    raise
                except Exception as table_err:
                    print(f"⚠️  Lỗi copy bảng: {table_err}")

            new_path = rename_output(src_path, goi_thau_id, used_names)

            if missing_placeholders:
                warn_msg = f"Warning: Placeholder {', '.join('{{' + k + '}}' for k in missing_placeholders)} không có data"
                results.append(f"⚠️ {tpl_name} → {new_path.name}\n   → {warn_msg}")
                logger.warning_count += 1
                logger.log_event("⚠️", f"{tpl_name} → {new_path.name}", warn_msg)
            else:
                results.append(f"✅ {tpl_name} → {new_path.name}")
                logger.ok_count += 1
                logger.log_event("✅", f"{tpl_name} → {new_path.name}")

        except PermissionError as e:
            if "being used by another process" in str(e) or getattr(e, 'errno', None) == errno.EACCES:
                has_locked = True
                failed_templates.append(tpl_name)
                err_msg = "Không thể ghi file vì đang được mở trong Word.\n   → Vui lòng đóng file này và nhấn \"🔄 Chạy lại file lỗi\""
                results.append(f"🔒 {tpl_name}: {err_msg}")
                logger.error_count += 1
                logger.log_event("🔒", tpl_name, f"Lỗi: PermissionError – {err_msg}")
            else:
                has_other_error = True
                failed_templates.append(tpl_name)
                err_msg = f"Lỗi quyền truy cập – {e}"
                results.append(f"❌ {tpl_name}: {err_msg}")
                logger.error_count += 1
                logger.log_event("❌", tpl_name, f"Lỗi: {err_msg}")

        except Exception as e:
            has_other_error = True
            failed_templates.append(tpl_name)
            err_msg = str(e)
            results.append(f"❌ {tpl_name}: {err_msg}")
            logger.error_count += 1
            logger.log_event("❌", tpl_name, f"Lỗi: {err_msg}")

        yield "\n".join(results), f"Đang xử lý {i + 1}/{total}...", None

    elapsed = time.time() - logger.start_time
    summary = f"✅ {logger.ok_count}  ⚠️ {logger.warning_count}  ❌ {logger.error_count}  /  {total} file  ({elapsed:.1f}s)"
    
    logger.write_footer()

    retry_state_data = None
    if failed_templates:
        retry_state_data = {
            "option_key": option_key,
            "package_label": package_label,
            "failed_templates": failed_templates,
            "all_locked": has_locked and not has_other_error,
        }

    yield "\n".join(results), summary, retry_state_data


def create_ui():
    init()

    _sel = {"opt": "", "pkg": "", "sheet_rows": [], "template_total": 0}

    with gr.Blocks(title=config.AppName) as app:
        gr.Markdown(f"# {config.AppName} – Xử lý tài liệu tự động")

        last_run_state = gr.State(None)

        with gr.Tabs() as tabs:
            with gr.Tab("1. Chọn & Chạy", id=0):
                with gr.Row():
                    with gr.Column(scale=1):
                        gr.Markdown("### Chọn Quy trình")
                        options = get_options()
                        option_radio = gr.Radio(choices=options, label="Chọn Quy trình")
                        package_radio = gr.Radio(choices=[], label="Chọn Dữ liệu")

                        with gr.Group():
                            gr.Markdown("**Preview thông tin:**")
                            pkg_preview = gr.Textbox(label="", interactive=False, max_lines=4)

                    with gr.Column(scale=1):
                        gr.Markdown("### Chọn file template & Chạy")
                        template_label = gr.Markdown("**Chọn template cần xử lý** (0 file)")
                        template_checkboxes = gr.CheckboxGroup(label="", choices=[])

                        with gr.Row():
                            select_all_btn = gr.Button("✓ Chọn tất cả")
                            deselect_all_btn = gr.Button("✗ Bỏ chọn tất cả")

                        run_btn = gr.Button("🚀 Chạy", variant="primary", size="lg")
                        check_btn = gr.Button("🔍 Kiểm tra", variant="secondary")
                        preview_box = gr.Textbox(
                            label="Kết quả kiểm tra",
                            interactive=False,
                            lines=4,
                            max_lines=8,
                            visible=False,
                        )

            with gr.Tab("2. Log & Kết quả", id=1):
                gr.Markdown("### Kết quả xử lý")
                result_log = gr.Textbox(
                    label="Chi tiết kết quả",
                    interactive=False,
                    lines=15,
                    max_lines=20,
                )
                status_text = gr.Textbox(label="Trạng thái", interactive=False)

                with gr.Row():
                    open_folder_btn = gr.Button("📂 Mở thư mục output", visible=False)
                    open_logs_btn = gr.Button("📋 Mở thư mục log", visible=True)
                    rerun_btn = gr.Button("← Chạy lại", variant="secondary")
                    retry_btn = gr.Button("🔄 Chạy lại file lỗi", variant="stop", visible=False)

        def on_package_change(pkg):
            opt = _sel["opt"]
            sheet_rows = _sel["sheet_rows"]
            details = get_package_details(opt, pkg, sheet_rows)
            if not details:
                preview_text = ""
            else:
                lines = [f"{k}: {v}" for k, v in details.items() if v]
                preview_text = "\n".join(lines)

            if not opt or not pkg:
                _sel["template_total"] = 0
                all_tpls = get_all_option_templates(opt)
                return preview_text, gr.update(choices=all_tpls, value=[]), gr.update(value="**Chọn template cần xử lý** (0 file)"), None, gr.update(visible=False), gr.update(visible=False)

            templates = get_workflow_templates(opt, pkg, sheet_rows)
            choices = [t.get("Name", "") for t in templates]
            _sel["template_total"] = len(choices)
            label_text = f"**Chọn template cần xử lý** ({len(choices)} file)"
            _sel["pkg"] = pkg

            return preview_text, gr.update(choices=choices, value=[]), gr.update(value=label_text), None, gr.update(visible=False), gr.update(visible=False)

        package_radio.change(
            fn=on_package_change,
            inputs=[package_radio],
            outputs=[pkg_preview, template_checkboxes, template_label, last_run_state, retry_btn, preview_box]
        )

        def on_option_change(opt):
            _sel["opt"] = opt or ""
            _sel["sheet_rows"] = []
            _sel["template_total"] = 0
            all_tpls = get_all_option_templates(opt)
            if not opt:
                pkgs = []
            else:
                opt_config = get_option_config(opt)
                sheet = opt_config.get("sheet", "GoiThau")
                show_format = opt_config.get("show", "")
                sql = resolve_sheet_query(sheet)
                try:
                    rows = ds.query(f"SELECT * FROM ({sql}) ORDER BY CAST(TT AS INTEGER)")
                except Exception:
                    try:
                        rows = ds.query(sql)
                    except Exception:
                        rows = []
                _sel["sheet_rows"] = rows
                pkgs = [label for r in rows if (label := safe_format(show_format, r))]
            return (
                gr.update(choices=pkgs, value=None),
                gr.update(choices=all_tpls, value=[]),
                gr.update(value="**Chọn template cần xử lý** (0 file)"),
                "",
                None,
                gr.update(visible=False),
                gr.update(visible=False),
            )

        option_radio.change(
            fn=on_option_change,
            inputs=[option_radio],
            outputs=[package_radio, template_checkboxes, template_label, pkg_preview, last_run_state, retry_btn, preview_box]
        )

        def update_checkbox_label(selected):
            total = _sel["template_total"]
            count = len(selected) if selected else 0
            return gr.update(value=f"**Chọn template cần xử lý** ({count}/{total} file)")

        template_checkboxes.change(fn=update_checkbox_label, inputs=[template_checkboxes], outputs=[template_label])

        def select_all():
            opt, pkg = _sel["opt"], _sel["pkg"]
            if not opt or not pkg:
                return gr.update(value=[])
            templates = get_workflow_templates(opt, pkg, _sel["sheet_rows"])
            choices = [t.get("Name", "") for t in templates]
            return gr.update(value=choices)

        def deselect_all():
            return gr.update(value=[])

        select_all_btn.click(fn=select_all, outputs=[template_checkboxes])
        deselect_all_btn.click(fn=deselect_all, outputs=[template_checkboxes])

        def on_check(opt, pkg, selected):
            result = run_preview(opt, pkg, selected)
            return gr.update(value=result, visible=True)

        check_btn.click(
            fn=on_check,
            inputs=[option_radio, package_radio, template_checkboxes],
            outputs=[preview_box],
        )

        def get_retry_label(retry_state):
            if not retry_state or not retry_state.get("failed_templates"):
                return gr.update(visible=False, interactive=True)
            n = len(retry_state["failed_templates"])
            if retry_state.get("all_locked"):
                return gr.update(visible=True, value=f"🔄 Chạy lại ({n} file – đã đóng file chưa?)", interactive=True)
            return gr.update(visible=True, value=f"🔄 Chạy lại {n} file lỗi", interactive=True)

        def disable_run():
            return gr.update(interactive=False)

        def enable_run():
            return gr.update(interactive=True)

        run_event = run_btn.click(
            fn=disable_run,
            outputs=[run_btn],
        ).then(
            fn=run_batch,
            inputs=[option_radio, package_radio, template_checkboxes],
            outputs=[result_log, status_text, last_run_state],
            show_progress="full",
            trigger_mode="once",
        ).then(
            lambda: gr.update(visible=True),
            outputs=[open_folder_btn]
        ).then(
            get_retry_label,
            inputs=[last_run_state],
            outputs=[retry_btn],
        ).then(
            enable_run,
            outputs=[run_btn],
        )

        def on_open_folder():
            try:
                open_output_folder(config)
                return "✅ Thư mục đã mở"
            except Exception as e:
                return f"❌ Lỗi mở thư mục: {e}"

        open_folder_btn.click(fn=on_open_folder, outputs=[status_text])

        def on_open_logs():
            try:
                open_logs_folder(config)
                return "✅ Thư mục log đã mở"
            except Exception as e:
                return f"❌ Lỗi mở thư mục log: {e}"

        open_logs_btn.click(fn=on_open_logs, outputs=[status_text])

        def on_retry_click(retry_state):
            if not retry_state or not retry_state.get("failed_templates"):
                return "⚠️ Không có file lỗi nào để chạy lại", None, gr.update(visible=False)
            n = len(retry_state["failed_templates"])
            if retry_state.get("all_locked"):
                return f"🔄 Đang chạy lại {n} file (đã đóng file chưa?)...", retry_state, gr.update(visible=True)
            return f"🔄 Đang chạy lại {n} file lỗi...", retry_state, gr.update(visible=True)

        async def run_retry_batch(retry_state, progress=gr.Progress()):
            if not retry_state or not retry_state.get("failed_templates"):
                yield "⚠️ Không có file lỗi nào để chạy lại", "⚠️ Không có file lỗi", None
                return
            option_key = retry_state["option_key"]
            package_label = retry_state["package_label"]
            failed_templates = retry_state["failed_templates"]
            async for log, status, new_state in run_batch(option_key, package_label, failed_templates, progress, retry_state):
                yield log, status, new_state

        def disable_retry():
            return gr.update(interactive=False)

        retry_event = retry_btn.click(
            fn=disable_retry,
            outputs=[retry_btn],
        ).then(
            fn=on_retry_click,
            inputs=[last_run_state],
            outputs=[status_text, last_run_state, retry_btn],
        ).then(
            fn=run_retry_batch,
            inputs=[last_run_state],
            outputs=[result_log, status_text, last_run_state],
            show_progress="full",
            trigger_mode="once",
        ).then(
            lambda: gr.update(visible=True),
            outputs=[open_folder_btn],
        ).then(
            get_retry_label,
            inputs=[last_run_state],
            outputs=[retry_btn],
        )

        def on_rerun():
            _sel["opt"] = ""
            _sel["pkg"] = ""
            _sel["sheet_rows"] = []
            _sel["template_total"] = 0
            initial_options = get_options()
            return (
                gr.update(choices=initial_options, value=None),
                gr.update(choices=[], value=None),
                gr.update(value=[]),
                gr.update(value=""),
                gr.update(visible=False),
                gr.update(visible=False),
                gr.update(value=""),
                gr.update(value=""),
                gr.update(selected=0),
                None,
            )

        rerun_btn.click(
            fn=on_rerun,
            outputs=[option_radio, package_radio, template_checkboxes, pkg_preview, open_folder_btn, retry_btn, result_log, status_text, tabs, last_run_state],
        )

    return app


if __name__ == "__main__":
    app = create_ui()
    PORT = 7864
    while True:
        if PORT > 7900:
            raise RuntimeError("Không tìm thấy port trống (7864–7900)")
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.bind(('127.0.0.1', PORT))
            sock.close()
            break
        except OSError:
            PORT += 1

    threading.Thread(target=lambda: webbrowser.open(f"http://127.0.0.1:{PORT}"), daemon=True).start()
    print(f"KisorDoc-AI running at http://127.0.0.1:{PORT}")
    app.launch(server_port=PORT, share=False, quiet=True, inbrowser=False)
