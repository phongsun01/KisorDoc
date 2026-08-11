"""
kisorlib/text_reporter.py
──────────────────────────
Xuất HTML và Excel report cho text_migrator (dry-run preview).

Tách riêng khỏi text_migrator.py để:
  - text_migrator không phụ thuộc openpyxl style (chỉ cần openpyxl core để đọc)
  - Dễ test độc lập
  - app.py có thể gọi trực tiếp nếu cần hiển thị report trong Gradio tab

Public API:
    generate_html_report(results, report_dir, *, excel_path, sample_row) -> Path
    generate_excel_report(results, report_dir) -> Path
"""

from __future__ import annotations

import datetime
import html as html_lib
import os
from pathlib import Path

from kisorlib.text_migrator import TextFileResult


# ──────────────────────────────────────────────────────────────────────────────
# HTML Report
# ──────────────────────────────────────────────────────────────────────────────

def generate_html_report(
    results:    list[TextFileResult],
    report_dir: Path | str,
    *,
    excel_path: str = "",
    sample_row: int = 0,
) -> Path:
    """
    Xuất file HTML preview dry-run.
    Trả về Path của file đã tạo.
    """
    report_dir = Path(report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)

    ts         = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path   = report_dir / f"dryrun_{ts}.html"
    now_str    = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    total_files   = len(results)
    changed_files = sum(1 for r in results if r.changed)
    total_changes = sum(len(r.changes) for r in results)

    # ── Build table rows ──────────────────────────────────────────────────────
    rows_html = ""
    for r in results:
        fname = html_lib.escape(r.path.name)
        warn_html = ""
        if r.warnings:
            warn_html = "<div class='warn'>" + "<br>".join(
                html_lib.escape(w) for w in r.warnings
            ) + "</div>"

        if not r.success:
            rows_html += f"""
    <tr class="err">
      <td class="fcell">{fname}<br><small style="color:#e74c3c">{html_lib.escape(r.error or '')}</small></td>
      <td colspan="3" style="color:#e74c3c">— Lỗi xử lý —</td>
    </tr>"""
        elif not r.changed:
            rows_html += f"""
    <tr class="nochg">
      <td class="fcell">{fname}{warn_html}</td>
      <td colspan="3" style="color:#aaa;font-style:italic">— Không có thay đổi —</td>
    </tr>"""
        else:
            # Nhóm changes theo paragraph để không lặp đoạn văn
            para_groups: dict[str, list] = {}
            for c in r.changes:
                para_groups.setdefault(c.paragraph, []).append(c)

            n_paras = len(para_groups)
            file_cell_written = False
            for para_text, para_changes in para_groups.items():
                fc = ""
                if not file_cell_written:
                    fc = (f'<td class="fcell" rowspan="{n_paras}">'
                          f'{fname}<br><small>{len(r.changes)} thay đổi</small>'
                          f'{warn_html}</td>')
                    file_cell_written = True

                details = "<br>".join(
                    f'<span class="del">{html_lib.escape(c.original)}</span>'
                    f' &rarr; '
                    f'<span class="ins">{html_lib.escape(c.placeholder)}</span>'
                    for c in para_changes
                )
                # Highlight placeholder trong preview
                preview = html_lib.escape(para_text)
                for c in para_changes:
                    preview = preview.replace(
                        html_lib.escape(c.placeholder),
                        f'<mark>{html_lib.escape(c.placeholder)}</mark>'
                    )

                rows_html += f"""
    <tr>
      {fc}
      <td class="para">{preview}</td>
      <td>{details}</td>
    </tr>"""

    # ── HTML template ─────────────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<title>Dry-run Preview — KisorDoc</title>
<style>
  *{{box-sizing:border-box}}
  body{{font-family:'Segoe UI',Arial,sans-serif;margin:0;background:#f5f6fa;color:#222}}
  header{{background:#1e3a5f;color:#fff;padding:18px 30px}}
  header h1{{margin:0 0 4px;font-size:1.2rem}}
  header p{{margin:0;font-size:.82rem;opacity:.75}}
  .stats{{display:flex;gap:16px;padding:14px 30px;background:#fff;border-bottom:1px solid #e0e0e0;flex-wrap:wrap}}
  .stat{{background:#f0f4ff;border-radius:8px;padding:9px 20px;text-align:center}}
  .stat .num{{font-size:1.5rem;font-weight:700;color:#1e3a5f}}
  .stat .lbl{{font-size:.72rem;color:#666}}
  .disclaimer{{margin:14px 30px 0;background:#fff8e1;border-left:4px solid #f0a500;
               padding:9px 14px;border-radius:0 6px 6px 0;font-size:.82rem;color:#7a5800}}
  .container{{padding:16px 30px 30px;overflow-x:auto}}
  table{{width:100%;border-collapse:collapse;background:#fff;
         box-shadow:0 1px 4px rgba(0,0,0,.08);border-radius:8px;overflow:hidden}}
  thead tr{{background:#1e3a5f;color:#fff}}
  thead th{{padding:10px 13px;text-align:left;font-size:.82rem;font-weight:600;white-space:nowrap}}
  tbody tr{{border-bottom:1px solid #eee}}
  tbody tr:hover{{background:#fafbff}}
  tbody tr.nochg{{background:#fafafa}}
  tbody tr.err{{background:#fff5f5}}
  td{{padding:8px 13px;vertical-align:top;font-size:.82rem;line-height:1.5}}
  td.fcell{{font-weight:600;color:#1e3a5f;white-space:nowrap;
             border-right:2px solid #e8edf5;min-width:160px}}
  td.fcell small{{font-weight:400;color:#888;display:block}}
  td.para{{max-width:360px;word-break:break-word}}
  mark{{background:#d4f5d4;color:#1a7a1a;border-radius:3px;padding:1px 3px;font-weight:600}}
  span.del{{color:#c0392b;text-decoration:line-through;background:#fde8e8;
             padding:1px 4px;border-radius:3px;font-weight:600}}
  span.ins{{color:#1a7a1a;background:#d4f5d4;padding:1px 4px;border-radius:3px;font-weight:600}}
  div.warn{{background:#fff8e1;border-left:3px solid #f0a500;padding:5px 8px;
             margin-top:5px;font-size:.78rem;color:#7a5800;border-radius:0 4px 4px 0}}
  footer{{text-align:center;padding:16px;font-size:.78rem;color:#aaa}}
</style>
</head>
<body>
<header>
  <h1>🔍 Dry-run Preview — KisorDoc Text Migrator</h1>
  <p>Excel: <b>{html_lib.escape(excel_path)}</b> &nbsp;|&nbsp;
     Row mẫu: <b>{sample_row}</b> &nbsp;|&nbsp;
     Tạo lúc: {now_str}</p>
</header>

<div class="stats">
  <div class="stat"><div class="num">{total_files}</div><div class="lbl">file .docx</div></div>
  <div class="stat"><div class="num">{changed_files}</div><div class="lbl">file có thay đổi</div></div>
  <div class="stat"><div class="num">{total_changes}</div><div class="lbl">lần thay thế</div></div>
</div>

<div class="disclaimer">
  ⚠ Preview hiển thị text ghép nối. Khi migrate thật, biên XML run có thể gây lệch đôi chút.
  Dùng <code>--verbose</code> để debug nếu nghi ngờ.
</div>

<div class="container">
<table>
  <thead>
    <tr>
      <th>File</th>
      <th>Đoạn văn (sau thay)</th>
      <th>Biến được thay</th>
    </tr>
  </thead>
  <tbody>
{rows_html}
  </tbody>
</table>
</div>
<footer>migrate_text_to_placeholders.py — kisorlib.text_reporter</footer>
</body>
</html>"""

    out_path.write_text(html, encoding="utf-8")
    return out_path


# ──────────────────────────────────────────────────────────────────────────────
# Excel Report
# ──────────────────────────────────────────────────────────────────────────────

def generate_excel_report(
    results:    list[TextFileResult],
    report_dir: Path | str,
) -> Path:
    """
    Xuất file Excel audit report.
    Trả về Path của file đã tạo.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError as e:
        raise ImportError("Cần cài openpyxl: pip install openpyxl") from e

    report_dir = Path(report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)

    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = report_dir / f"dryrun_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Migration Audit"

    # ── Styles ────────────────────────────────────────────────────────────────
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
    reg_font   = Font(name="Segoe UI", size=10)
    red_fill   = PatternFill("solid", fgColor="FFE8E8")
    green_fill = PatternFill("solid", fgColor="E8FFE8")
    grey_fill  = PatternFill("solid", fgColor="F7F7F7")
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center     = Alignment(vertical="center", horizontal="center")

    headers = ["File", "Đoạn văn (Preview)", "Nguyên bản mẫu", "Placeholder thay thế", "Cảnh báo"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill   = hdr_fill
        c.font   = hdr_font
        c.border = border
        c.alignment = center
    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    row_num = 2
    for r in results:
        fname    = r.path.name
        warn_str = " | ".join(r.warnings) if r.warnings else ""

        if not r.success:
            ws.cell(row=row_num, column=1, value=fname)
            ws.cell(row=row_num, column=2, value=f"[Lỗi] {r.error}")
            ws.cell(row=row_num, column=5, value=warn_str)
            for col in range(1, 6):
                c = ws.cell(row=row_num, column=col)
                c.font = reg_font; c.border = border; c.alignment = wrap_align
            row_num += 1
            continue

        if not r.changed:
            ws.cell(row=row_num, column=1, value=fname)
            c = ws.cell(row=row_num, column=2, value="— Không có thay đổi —")
            c.font = Font(name="Segoe UI", size=10, color="AAAAAA", italic=True)
            ws.cell(row=row_num, column=5, value=warn_str)
            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = grey_fill; cell.border = border; cell.alignment = wrap_align
            row_num += 1
            continue

        # Nhóm theo paragraph
        para_groups: dict[str, list] = {}
        for ch in r.changes:
            para_groups.setdefault(ch.paragraph, []).append(ch)

        first = True
        for para_text, para_changes in para_groups.items():
            for ch in para_changes:
                ws.cell(row=row_num, column=1, value=fname if first else "")
                c_para = ws.cell(row=row_num, column=2, value=para_text)
                c_orig = ws.cell(row=row_num, column=3, value=ch.original)
                c_ph   = ws.cell(row=row_num, column=4, value=ch.placeholder)
                ws.cell(row=row_num, column=5, value=warn_str if first else "")

                c_para.fill = red_fill
                c_ph.fill   = green_fill
                for col in range(1, 6):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = reg_font; cell.border = border; cell.alignment = wrap_align
                ws.row_dimensions[row_num].height = 40
                first = False
                row_num += 1

    # ── Column widths, freeze, filter ─────────────────────────────────────────
    col_widths = {"A": 26, "B": 50, "C": 28, "D": 28, "E": 36}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{row_num - 1}"

    wb.save(out_path)
    return out_path
