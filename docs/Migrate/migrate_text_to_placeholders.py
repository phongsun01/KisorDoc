"""
migrate_text_to_placeholders.py
================================
Chuyển giá trị mẫu cụ thể trong docx sang Jinja2 placeholder {{ TenBien }}.

Cách dùng:
    # Dry-run: xem trước + xuất HTML & Excel report
    python migrate_text_to_placeholders.py \
        --excel "data.xlsx" --row 1 --docx-dir "2. Templates/" --dry-run

    # Chạy thật (tự backup .bak.docx)
    python migrate_text_to_placeholders.py \
        --excel "data.xlsx" --row 1 --docx-dir "2. Templates/"

Tham số:
    --excel       File Excel chứa dữ liệu mẫu
    --row         Row mẫu 1-based (không tính header). VD: --row 1 = row đầu tiên data
    --docx-dir    Thư mục chứa các file .docx cần migrate
    --sheet       Tên sheet (mặc định: sheet đầu tiên)
    --dry-run     Không sửa file, chỉ xuất report HTML + Excel
    --report-dir  Thư mục lưu report (mặc định: thư mục hiện tại)
"""

import os
import sys
import zipfile
import shutil
import argparse
import html as html_lib
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("Thiếu thư viện: pip install openpyxl")
    sys.exit(1)

try:
    from lxml import etree
except ImportError:
    print("Thiếu thư viện: pip install lxml")
    sys.exit(1)


# ─── Namespace Word XML ────────────────────────────────────────────────────────
W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
W = f"{{{W_NS}}}"


# ─── Đọc mapping từ Excel ─────────────────────────────────────────────────────
def load_mapping(excel_path: str, sample_row_index: int, sheet_name: str = None) -> dict:
    """Trả về dict sorted by length desc: { "giá trị mẫu": "{{ TenBien }}" }"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Excel cần ít nhất 2 row (header + data).")
    header_row = rows[0]
    if sample_row_index >= len(rows):
        raise ValueError(
            f"Row {sample_row_index} không tồn tại. Excel chỉ có {len(rows)-1} row data."
        )
    sample_row = rows[sample_row_index]
    mapping = {}
    for col_name, cell_value in zip(header_row, sample_row):
        if col_name is None or cell_value is None:
            continue
        key = str(cell_value).strip()
        var_name = str(col_name).strip()
        if key and var_name:
            mapping[key] = f"{{{{ {var_name} }}}}"
    return dict(sorted(mapping.items(), key=lambda x: len(x[0]), reverse=True))


# ─── Xử lý XML ────────────────────────────────────────────────────────────────
def get_paragraph_text(para_elem) -> str:
    return "".join(t.text or "" for t in para_elem.findall(f".//{W}t"))


def replace_in_paragraph(para_elem, mapping: dict):
    """Thay thế và trả về (text_gốc, text_mới). Nếu không đổi trả về ('', '')."""
    full_text = get_paragraph_text(para_elem)
    new_text = full_text
    for value, placeholder in mapping.items():
        new_text = new_text.replace(value, placeholder)
    if new_text == full_text:
        return "", ""

    runs = para_elem.findall(f".//{W}r")
    if not runs:
        return "", ""

    first_run = runs[0]
    first_t = first_run.find(f"{W}t")
    if first_t is None:
        first_t = etree.SubElement(first_run, f"{W}t")
    first_t.text = new_text
    first_t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    for run in runs[1:]:
        t_elem = run.find(f"{W}t")
        if t_elem is not None:
            t_elem.text = ""
    return full_text, new_text


def process_xml(xml_bytes: bytes, mapping: dict, dry_run: bool):
    """Xử lý một file XML trong docx. Trả về (bytes_mới, list[dict thay đổi])."""
    changes = []
    tree = etree.fromstring(xml_bytes)

    for para in tree.findall(f".//{W}p"):
        original_text = get_paragraph_text(para)
        if not any(val in original_text for val in mapping):
            continue

        matches = [(val, ph) for val, ph in mapping.items() if val in original_text]

        if dry_run:
            preview = original_text
            for val, ph in mapping.items():
                preview = preview.replace(val, ph)
            changes.append({"original": original_text, "replaced": preview, "matches": matches})
        else:
            orig, new = replace_in_paragraph(para, mapping)
            if orig:
                changes.append({"original": orig, "replaced": new, "matches": matches})

    if dry_run:
        return xml_bytes, changes

    new_xml = etree.tostring(tree, xml_declaration=True, encoding="UTF-8", standalone=True)
    return new_xml, changes


def migrate_docx(filepath: str, mapping: dict, dry_run: bool) -> list:
    """Xử lý một file .docx. Trả về list thay đổi."""
    all_changes = []

    if not dry_run:
        backup_path = filepath.replace(".docx", ".bak.docx")
        if not os.path.exists(backup_path):
            shutil.copy2(filepath, backup_path)

    temp_path = filepath + ".tmp"
    try:
        with zipfile.ZipFile(filepath, "r") as zin:
            with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.startswith("word/") and item.filename.endswith(".xml"):
                        new_data, changes = process_xml(data, mapping, dry_run)
                        all_changes.extend(changes)
                        zout.writestr(item, new_data)
                    else:
                        zout.writestr(item, data)

        if not dry_run and all_changes:
            os.remove(filepath)
            os.rename(temp_path, filepath)
        elif os.path.exists(temp_path):
            os.remove(temp_path)

    except Exception as e:
        print(f"  ✗ Lỗi xử lý {os.path.basename(filepath)}: {e}")
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return all_changes


# ─── HTML Report ──────────────────────────────────────────────────────────────
def highlight_original(text: str, matches: list) -> str:
    escaped = html_lib.escape(text)
    for val, _ in sorted(matches, key=lambda x: len(x[0]), reverse=True):
        escaped = escaped.replace(
            html_lib.escape(val),
            f'<mark class="old">{html_lib.escape(val)}</mark>'
        )
    return escaped


def highlight_replaced(text: str, matches: list) -> str:
    escaped = html_lib.escape(text)
    for _, ph in matches:
        escaped = escaped.replace(
            html_lib.escape(ph),
            f'<mark class="new">{html_lib.escape(ph)}</mark>'
        )
    return escaped


def export_html(report_rows: list, out_path: str, excel_path: str, sample_row: int):
    now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    total_changes = sum(r["change_count"] for r in report_rows)
    files_changed = sum(1 for r in report_rows if r["change_count"] > 0)

    rows_html_parts = []
    for r in report_rows:
        fname = html_lib.escape(r["file"])
        if r["change_count"] == 0:
            rows_html_parts.append(f"""
    <tr class="no-change">
      <td class="file-cell">{fname}<br><small>0 thay đổi</small></td>
      <td colspan="3" style="color:#aaa;font-style:italic">— Không có thay đổi —</td>
    </tr>""")
        else:
            for i, ch in enumerate(r["changes"]):
                file_cell = (
                    f'<td class="file-cell" rowspan="{r["change_count"]}">'
                    f'{fname}<br><small>{r["change_count"]} thay đổi</small></td>'
                ) if i == 0 else ""

                orig_html  = highlight_original(ch["original"], ch["matches"])
                repl_html  = highlight_replaced(ch["replaced"], ch["matches"])
                tags_html  = "".join(
                    f'<span class="tag">{html_lib.escape(v)} → {html_lib.escape(p)}</span>'
                    for v, p in ch["matches"]
                )
                rows_html_parts.append(f"""
    <tr>
      {file_cell}
      <td class="para">{orig_html}</td>
      <td class="para">{repl_html}</td>
      <td>{tags_html}</td>
    </tr>""")

    rows_html = "\n".join(rows_html_parts)

    html = f"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<title>Dry-run Preview Report</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; background: #f5f6fa; color: #222; }}
  header {{ background: #1e3a5f; color: #fff; padding: 20px 32px; }}
  header h1 {{ margin: 0 0 4px; font-size: 1.3rem; }}
  header p {{ margin: 0; font-size: .85rem; opacity: .8; }}
  .stats {{ display: flex; gap: 20px; padding: 16px 32px; background: #fff; border-bottom: 1px solid #e0e0e0; flex-wrap: wrap; }}
  .stat {{ background: #f0f4ff; border-radius: 8px; padding: 10px 24px; text-align: center; }}
  .stat .num {{ font-size: 1.6rem; font-weight: 700; color: #1e3a5f; }}
  .stat .lbl {{ font-size: .75rem; color: #666; }}
  .container {{ padding: 20px 32px; overflow-x: auto; }}
  table {{ width: 100%; border-collapse: collapse; background: #fff;
           box-shadow: 0 1px 4px rgba(0,0,0,.08); border-radius: 8px; overflow: hidden; }}
  thead tr {{ background: #1e3a5f; color: #fff; }}
  thead th {{ padding: 11px 14px; text-align: left; font-size: .85rem; font-weight: 600; white-space: nowrap; }}
  tbody tr {{ border-bottom: 1px solid #eee; }}
  tbody tr:hover {{ background: #fafbff; }}
  tbody tr.no-change {{ background: #fafafa; }}
  td {{ padding: 9px 14px; vertical-align: top; font-size: .85rem; line-height: 1.55; }}
  td.file-cell {{ font-weight: 600; color: #1e3a5f; white-space: nowrap;
                  border-right: 2px solid #e8edf5; min-width: 170px; }}
  td.file-cell small {{ font-weight: 400; color: #888; display: block; }}
  td.para {{ max-width: 340px; word-break: break-word; }}
  mark.old {{ background: #ffe0e0; color: #c0392b; border-radius: 3px;
              padding: 1px 3px; font-weight: 600; text-decoration: line-through; }}
  mark.new {{ background: #d4f5d4; color: #1a7a1a; border-radius: 3px;
              padding: 1px 3px; font-weight: 600; }}
  span.tag {{ display: inline-block; background: #eef2ff; color: #3a4db7;
              border: 1px solid #c5cdf7; border-radius: 4px; padding: 2px 7px;
              font-size: .78rem; margin: 2px 2px 2px 0; white-space: nowrap; }}
  footer {{ text-align: center; padding: 20px; font-size: .8rem; color: #aaa; }}
</style>
</head>
<body>
<header>
  <h1>🔍 Dry-run Preview Report</h1>
  <p>Excel: <b>{html_lib.escape(excel_path)}</b> &nbsp;|&nbsp; Row mẫu: <b>{sample_row}</b> &nbsp;|&nbsp; Tạo lúc: {now}</p>
</header>
<div class="stats">
  <div class="stat"><div class="num">{len(report_rows)}</div><div class="lbl">file .docx</div></div>
  <div class="stat"><div class="num">{files_changed}</div><div class="lbl">file có thay đổi</div></div>
  <div class="stat"><div class="num">{total_changes}</div><div class="lbl">đoạn văn sẽ đổi</div></div>
</div>
<div class="container">
<table>
  <thead>
    <tr>
      <th>File</th>
      <th>Đoạn văn gốc</th>
      <th>Sau khi thay</th>
      <th>Biến được thay</th>
    </tr>
  </thead>
  <tbody>
{rows_html}
  </tbody>
</table>
</div>
<footer>migrate_text_to_placeholders.py — dry-run preview</footer>
</body>
</html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


# ─── Excel Report ─────────────────────────────────────────────────────────────
def export_excel_report(report_rows: list, out_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dry-run Report"

    headers = ["File", "Đoạn văn gốc", "Sau khi thay", "Biến được thay", "Số thay đổi"]
    hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    old_fill   = PatternFill("solid", fgColor="FFE8E8")
    new_fill   = PatternFill("solid", fgColor="E8FFE8")
    none_fill  = PatternFill("solid", fgColor="F7F7F7")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 22

    row_num = 2
    for r in report_rows:
        if r["change_count"] == 0:
            ws.cell(row=row_num, column=1, value=r["file"])
            c = ws.cell(row=row_num, column=2, value="— Không có thay đổi —")
            c.font = Font(color="AAAAAA", italic=True)
            ws.cell(row=row_num, column=5, value=0)
            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = none_fill
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_num += 1
        else:
            for i, ch in enumerate(r["changes"]):
                match_str = "  |  ".join(f"{v}  →  {p}" for v, p in ch["matches"])
                ws.cell(row=row_num, column=1, value=r["file"] if i == 0 else "")
                c_orig = ws.cell(row=row_num, column=2, value=ch["original"])
                c_repl = ws.cell(row=row_num, column=3, value=ch["replaced"])
                ws.cell(row=row_num, column=4, value=match_str)
                ws.cell(row=row_num, column=5, value=r["change_count"] if i == 0 else "")
                c_orig.fill = old_fill
                c_repl.fill = new_fill
                for col in range(1, 6):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                ws.row_dimensions[row_num].height = 42
                row_num += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{row_num - 1}"

    wb.save(out_path)


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Migrate giá trị mẫu trong docx sang Jinja2 placeholder"
    )
    parser.add_argument("--excel",      required=True, help="Đường dẫn file Excel")
    parser.add_argument("--row",        type=int, required=True,
                        help="Row mẫu 1-based (không tính header). VD: --row 1 = row đầu tiên data")
    parser.add_argument("--docx-dir",   required=True, help="Thư mục chứa file .docx")
    parser.add_argument("--sheet",      default=None,  help="Tên sheet (mặc định: sheet đầu tiên)")
    parser.add_argument("--dry-run",    action="store_true",
                        help="Xem trước, xuất HTML + Excel report, KHÔNG sửa file")
    parser.add_argument("--report-dir", default=".",
                        help="Thư mục lưu report (mặc định: thư mục hiện tại)")
    args = parser.parse_args()

    # 1. Load mapping
    print(f"\n📊 Đọc Excel: {args.excel}  (row {args.row})")
    mapping = load_mapping(args.excel, args.row, args.sheet)
    print(f"   {len(mapping)} biến tìm thấy:")
    for val, ph in list(mapping.items())[:8]:
        print(f'   "{val}"  →  {ph}')
    if len(mapping) > 8:
        print(f"   ... và {len(mapping) - 8} biến khác")

    if not mapping:
        print("⚠️  Không có mapping nào. Kiểm tra lại file Excel.")
        return

    mode = "🔍 DRY RUN — không sửa file" if args.dry_run else "⚙️  MIGRATE — sẽ backup .bak.docx"
    print(f"\n{mode}")
    print(f"📁 Thư mục: {args.docx_dir}\n")

    # 2. Duyệt file
    report_rows = []
    for root, _, files in os.walk(args.docx_dir):
        for fname in sorted(files):
            if not fname.endswith(".docx") or fname.endswith(".bak.docx"):
                continue
            fpath = os.path.join(root, fname)
            changes = migrate_docx(fpath, mapping, dry_run=args.dry_run)
            report_rows.append({"file": fname, "changes": changes, "change_count": len(changes)})
            tag = "[DRY]" if args.dry_run else "[OK] "
            status = f"✓ {len(changes)} đoạn" if changes else "— không có thay đổi"
            print(f"  {tag} {fname}  {status}")

    total_changes = sum(r["change_count"] for r in report_rows)
    files_changed = sum(1 for r in report_rows if r["change_count"] > 0)
    print(f"\n{'📋' if args.dry_run else '✅'} {files_changed}/{len(report_rows)} file | {total_changes} đoạn văn sẽ thay đổi")

    # 3. Xuất report khi dry-run
    if args.dry_run:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        os.makedirs(args.report_dir, exist_ok=True)

        html_path = os.path.join(args.report_dir, f"dryrun_{ts}.html")
        export_html(report_rows, html_path, args.excel, args.row)
        print(f"\n  📄 HTML  → {html_path}")

        xlsx_path = os.path.join(args.report_dir, f"dryrun_{ts}.xlsx")
        export_excel_report(report_rows, xlsx_path)
        print(f"  📊 Excel → {xlsx_path}")

        print("\n  ✔ Kiểm tra report xong, bỏ --dry-run để chạy thật.\n")


if __name__ == "__main__":
    main()
