"""
PATCH: dataset.py
BUG FIX: Mở file Excel 2 lần mỗi sheet → gộp lại thành 1 lần đọc trong _load()
"""

import duckdb
import pandas as pd
import openpyxl
from pathlib import Path
from config import AppConfig


class DataSet:
    def __init__(self, config: AppConfig):
        self.config = config
        self.conn = duckdb.connect(":memory:")
        self.table_names: list[str] = []
        self._load()

    def _load(self):
        exception_prefix = self.config.ExceptionSheet
        for xlsx_file in sorted(self.config.data_path.glob("*.xlsx")):
            if xlsx_file.name.startswith("~$"):
                continue
            # FIX: Chỉ mở file 1 lần duy nhất, dùng data_only=True luôn
            try:
                wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
            except Exception:
                continue

            for sheet_name in wb.sheetnames:
                stripped = sheet_name.strip()
                if stripped.startswith(exception_prefix):
                    continue
                safe_name = _safe_table_name(stripped)
                if safe_name in self.table_names:
                    continue

                # FIX: Đọc data trực tiếp từ wb đang mở, không mở lại
                df = _ws_to_dataframe(wb[sheet_name])
                if df is not None and not df.empty:
                    self.conn.register(safe_name, df)
                    self.table_names.append(safe_name)

            wb.close()

    def query(self, sql: str) -> list[dict]:
        return self.conn.execute(sql).fetchdf().to_dict(orient="records")

    def get_table(self, name: str) -> pd.DataFrame | None:
        safe = _safe_table_name(name)
        if safe in self.table_names:
            return self.conn.execute(f'SELECT * FROM "{safe}"').fetchdf()
        return None

    def close(self):
        self.conn.close()


def _ws_to_dataframe(ws) -> pd.DataFrame | None:
    """
    FIX: Nhận worksheet object trực tiếp thay vì mở lại file.
    Dùng chung cho cả _load() — không cần _sheet_to_dataframe() riêng nữa.
    """
    try:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None
        header = [
            str(c) if c is not None else f"Col{i}"
            for i, c in enumerate(rows[0])
        ]
        data_rows = [row for row in rows[1:] if any(v is not None for v in row)]
        if not data_rows:
            return None
        df = pd.DataFrame(data_rows, columns=header)
        df = df.dropna(how="all")
        return df
    except Exception:
        return None


def _safe_table_name(name: str) -> str:
    safe = "".join(c if c.isalnum() or c == "_" else "_" for c in name)
    if not safe or safe[0].isdigit():
        safe = "t_" + safe
    return safe
