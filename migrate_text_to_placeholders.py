#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import re
import argparse
import shutil
import zipfile
import datetime
import unicodedata
from openpyxl import load_workbook
from lxml import etree

# Import clean_config_key từ kisorlib.utils
try:
    from kisorlib.utils import clean_config_key
except ImportError:
    # Fallback dự phòng nếu chạy bên ngoài môi trường kisorlib
    def clean_config_key(key: str) -> str:
        clean = key.strip("<>{}| ")
        for suffix in (".Date.Long", ".Date.long", ".date_long"):
            if clean.endswith(suffix):
                return clean[:-len(suffix)] + "_Date"
        if clean.endswith(".Date") or clean.endswith(".date"):
            return clean[:-5] + "_Date"
        if clean.endswith(".Day") or clean.endswith(".day"):
            return clean[:-4] + "_Date"
        if clean.endswith(".Month") or clean.endswith(".month"):
            return clean[:-6] + "_Date"
        if clean.endswith(".Year") or clean.endswith(".year"):
            return clean[:-5] + "_Date"
        for suffix in (".Upper", ".upper", ".Number", ".number"):
            if clean.endswith(suffix):
                clean = clean[:-len(suffix)]
                break
        if "|" in clean:
            clean = clean.split("|")[0].strip()
        return clean

# XML Namespaces
NAMESPACES = {
    'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
}

def ns_tag(tag):
    prefix, local = tag.split(':')
    return f"{{{NAMESPACES[prefix]}}}{local}"

def to_slug(text):
    """
    Sinh slug tiếng Việt không dấu, khoảng trắng đổi thành gạch dưới, giữ nguyên chữ hoa/thường
    Ví dụ: 'Họ và Tên' -> 'Ho_va_Ten'
    """
    if not text:
        return ""
    text_str = str(text)
    # Thay thế đ và Đ thủ công trước khi normalize
    text_str = text_str.replace('đ', 'd').replace('Đ', 'D')
    # Loại bỏ dấu tiếng Việt
    normalized = unicodedata.normalize('NFKD', text_str)
    no_accents = normalized.encode('ASCII', 'ignore').decode('utf-8')
    # Thay thế ký tự không phải chữ/số thành gạch dưới, nén nhiều gạch dưới liên tiếp
    slug = re.sub(r'[^a-zA-Z0-9]', '_', no_accents)
    slug = re.sub(r'_+', '_', slug)
    return slug.strip('_')

def load_mapping(excel_path, row_idx, sheet_name=None, config_sheet_name=None, min_length=3, verbose=False):
    """
    Đọc Excel và Config sheet để build mapping.
    """
    wb = load_workbook(excel_path, data_only=True)
    
    # 1. Đọc sheet dữ liệu
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]
        
    headers = [cell.value for cell in ws[1]]
    # row_idx là 1-based index (bỏ qua header) => row thực tế = row_idx + 1
    target_row = ws[row_idx + 1]
    row_values = [cell.value for cell in target_row]
    
    # 2. Đọc Config sheet (nếu có)
    config_map = {}
    config_sheet = config_sheet_name or "Config"
    if config_sheet in wb.sheetnames:
        c_ws = wb[config_sheet]
        # Tìm cột Key và Value
        c_headers = [cell.value for cell in c_ws[1]]
        try:
            key_idx = c_headers.index("Key")
            val_idx = c_headers.index("Value")
            for r in range(2, c_ws.max_row + 1):
                k_val = c_ws.cell(row=r, column=key_idx+1).value
                v_val = c_ws.cell(row=r, column=val_idx+1).value
                if k_val and v_val:
                    # Normalize Config.Value: strip và lower case
                    norm_val = str(v_val).strip().lower()
                    config_map[norm_val] = str(k_val).strip()
        except ValueError:
            if verbose:
                print(f"[Warning] Config sheet '{config_sheet}' must contain 'Key' and 'Value' headers.")
    elif verbose and config_sheet_name:
        print(f"[Warning] Config sheet '{config_sheet}' not found.")

    # 3. Kết hợp tạo cặp mapping
    mapping = {}
    collisions = []
    
    for h, val in zip(headers, row_values):
        if val is None or h is None:
            continue
        val_str = str(val).strip()
        if len(val_str) < min_length:
            continue
            
        h_norm = str(h).strip().lower()
        if h_norm in config_map:
            key = config_map[h_norm]
        else:
            key = to_slug(h)
            
        key_cleaned = clean_config_key(key)
        placeholder = f"{{{{ {key_cleaned} }}}}"
        
        if val_str in mapping:
            collisions.append((val_str, mapping[val_str], placeholder))
            if verbose:
                print(f"[Collision Warning] '{val_str}' mapped to '{mapping[val_str]}' first. Ignoring '{placeholder}'.")
        else:
            mapping[val_str] = placeholder
            
    # Sắp xếp theo chiều dài giảm dần của giá trị mẫu
    sorted_mapping = sorted(mapping.items(), key=lambda x: len(x[0]), reverse=True)
    return sorted_mapping, collisions

def get_xml_files(docx_path):
    """Trích xuất danh sách file XML có thể chứa text trong docx."""
    with zipfile.ZipFile(docx_path, 'r') as z:
        files = z.namelist()
    targets = []
    for f in files:
        if f.startswith('word/') and f.endswith('.xml'):
            # Lọc các file chính, header, footer, footnote, endnote
            if 'document' in f or 'header' in f or 'footer' in f or 'footnote' in f or 'endnote' in f:
                targets.append(f)
    return targets

def reconstruct_paragraph_text(p_node):
    """
    Ghép tất cả text từ các run trong paragraph và sinh offset_map.
    offset_map[char_index] = (run_node, char_offset_in_run_text, original_text_node)
    """
    full_text = ""
    offset_map = []
    
    # Tìm tất cả w:r hoặc w:ins (đôi khi track changes dùng w:ins chứa w:r)
    # Lấy w:t bên dưới chúng
    t_nodes = p_node.xpath('.//w:t', namespaces=NAMESPACES)
    
    for t_node in t_nodes:
        # Tìm run cha (w:r)
        r_node = t_node.getparent()
        t_text = t_node.text or ""
        for i, char in enumerate(t_text):
            offset_map.append((r_node, i, t_node))
        full_text += t_text
        
    return full_text, offset_map

def find_forbidden_spans(full_text):
    """Tìm các vùng cấm sửa (nằm trong {{ ... }})"""
    forbidden = []
    for m in re.finditer(r'\{\{.*?\}\}', full_text):
        forbidden.append(m.span())
    return forbidden

def is_overlap(span1, span2):
    return not (span1[1] <= span2[0] or span2[1] <= span1[0])

def safe_replace_xml(p_node, sorted_mapping, case_insensitive, dense_threshold, file_logs, file_path):
    """
    Thực hiện tìm kiếm và thay thế an toàn trên XML paragraph.
    """
    full_text, offset_map = reconstruct_paragraph_text(p_node)
    if not full_text:
        return False
        
    forbidden_spans = find_forbidden_spans(full_text)
    modified = False
    
    # Biến theo dõi tần suất thay thế từng token để phát hiện Dense Match
    p_matches = []
    
    # Duyệt qua các mapping đã sort by length
    for sample, placeholder in sorted_mapping:
        pattern = re.escape(sample)
        flags = re.IGNORECASE if case_insensitive else 0
        
        # Tìm tất cả matches
        matches = list(re.finditer(pattern, full_text, flags))
        if not matches:
            continue
            
        # Thao tác từng match từ phải qua trái (ngược chiều) để không làm lệch offset_map
        for m in reversed(matches):
            m_span = m.span()
            # Check idempotent
            overlap = False
            for f_span in forbidden_spans:
                if is_overlap(m_span, f_span):
                    overlap = True
                    break
            if overlap:
                continue
                
            # Xác định các run liên quan
            start_idx, end_idx = m_span[0], m_span[1]
            start_info = offset_map[start_idx]
            end_info = offset_map[end_idx - 1]
            
            start_run, start_offset, start_t = start_info
            end_run, end_offset, end_t = end_info
            
            # Ghi nhận log match
            p_matches.append({
                'sample': sample,
                'placeholder': placeholder,
                'original': full_text[start_idx:end_idx]
            })
            
            # Xử lý thay thế
            if start_t == end_t:
                # Text nằm gọn trong 1 thẻ w:t
                orig_text = start_t.text
                new_text = orig_text[:start_offset] + placeholder + orig_text[start_offset + len(sample):]
                start_t.text = new_text
            else:
                # Text vắt ngang nhiều run/w:t
                # 1. Start run: sửa text giữ phần tiền tố + placeholder
                orig_start_text = start_t.text
                start_t.text = orig_start_text[:start_offset] + placeholder
                
                # 2. Xóa các w:t ở giữa
                # Tìm tất cả w:t nằm giữa start_t và end_t trong paragraph
                all_t = p_node.xpath('.//w:t', namespaces=NAMESPACES)
                try:
                    s_idx = all_t.index(start_t)
                    e_idx = all_t.index(end_t)
                    for idx in range(s_idx + 1, e_idx):
                        all_t[idx].text = "" # Xóa nội dung
                except ValueError:
                    pass
                    
                # 3. End run: trim phần trùng, giữ hậu tố
                orig_end_text = end_t.text
                # trim phần trùng (ở đầu của end_t)
                end_t.text = orig_end_text[end_offset + 1:]
                
            modified = True
            # Reconstruct lại để update offset_map và forbidden_spans cho các mapping tiếp theo
            full_text, offset_map = reconstruct_paragraph_text(p_node)
            forbidden_spans = find_forbidden_spans(full_text)
            
    if modified:
        file_logs.append({
            'original_p': full_text, # text sau khi thay thế để preview
            'matches': p_matches
        })
        
    return modified

def process_docx(docx_path, sorted_mapping, case_insensitive, dense_threshold, dry_run, verbose):
    """
    Xử lý một file docx đơn lẻ.
    """
    xml_files = get_xml_files(docx_path)
    file_logs = []
    modified_any = False
    
    # Đọc XML từ ZIP
    xml_data = {}
    with zipfile.ZipFile(docx_path, 'r') as z:
        for x_file in xml_files:
            xml_data[x_file] = z.read(x_file)
            
    # Thao tác từng file XML
    for x_file, data in xml_data.items():
        root = etree.fromstring(data)
        xml_modified = False
        
        # Lấy tất cả paragraph w:p
        paragraphs = root.xpath('//w:p', namespaces=NAMESPACES)
        for p in paragraphs:
            # Thực hiện an toàn
            if safe_replace_xml(p, sorted_mapping, case_insensitive, dense_threshold, file_logs, docx_path):
                xml_modified = True
                
        if xml_modified:
            modified_any = True
            xml_data[x_file] = etree.tostring(root, encoding='utf-8', xml_declaration=True)
            
    # Cảnh báo Dense Match nếu vượt ngưỡng
    dense_warnings = []
    match_counts = {}
    for log in file_logs:
        for m in log['matches']:
            match_counts[m['sample']] = match_counts.get(m['sample'], 0) + 1
            
    for sample, count in match_counts.items():
        if count > dense_threshold:
            w_msg = f"[Dense Match Warning] '{sample}' matched {count} times (threshold: {dense_threshold})"
            dense_warnings.append(w_msg)
            if verbose:
                print(w_msg)
                
    if modified_any and not dry_run:
        # Ghi đè thực tế
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = docx_path.replace(".docx", f".{timestamp}.bak.docx")
        shutil.copy2(docx_path, backup_path)
        
        temp_path = docx_path + ".tmp"
        try:
            with zipfile.ZipFile(docx_path, 'r') as yin:
                with zipfile.ZipFile(temp_path, 'w') as yout:
                    for item in yin.infolist():
                        if item.filename in xml_data:
                            yout.writestr(item.filename, xml_data[item.filename])
                        else:
                            yout.writestr(item, yin.read(item.filename))
            os.remove(docx_path)
            os.rename(temp_path, docx_path)
        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise e
            
    return modified_any, file_logs, dense_warnings

def generate_html_report(report_dir, run_logs, timestamp):
    """Sinh báo cáo HTML trực quan."""
    os.makedirs(report_dir, exist_ok=True)
    report_path = os.path.join(report_dir, f"dryrun_{timestamp}.html")
    
    # Phép đếm thống kê
    total_files = len(run_logs)
    modified_files = sum(1 for x in run_logs.values() if x['modified'])
    total_changes = sum(len(x['logs']) for x in run_logs.values())
    
    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Dry-Run Migration Report</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 20px; background-color: #f8f9fa; }}
        h1 {{ color: #2c3e50; }}
        .summary-box {{ display: flex; gap: 20px; margin-bottom: 20px; }}
        .card {{ background: #fff; padding: 15px 25px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); flex: 1; }}
        .card h3 {{ margin: 0; color: #7f8c8d; font-size: 14px; text-transform: uppercase; }}
        .card p {{ margin: 5px 0 0 0; font-size: 24px; font-weight: bold; color: #2c3e50; }}
        .disclaimer {{ background: #fff3cd; color: #856404; padding: 12px; border-radius: 6px; border: 1px solid #ffeeba; margin-bottom: 20px; font-size: 14px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 10px; background: #fff; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }}
        th, td {{ padding: 12px; border: 1px solid #e0e0e0; text-align: left; }}
        th {{ background-color: #f1f2f6; color: #2c3e50; }}
        .del {{ color: #c0392b; text-decoration: line-through; background-color: #fde8e8; padding: 2px 4px; border-radius: 4px; }}
        .ins {{ color: #27ae60; font-weight: bold; background-color: #e8f8f0; padding: 2px 4px; border-radius: 4px; }}
        .warning-box {{ background-color: #fdf2e9; border-left: 4px solid #e67e22; padding: 10px; margin: 10px 0; border-radius: 0 4px 4px 0; font-size: 13px; }}
    </style>
</head>
<body>
    <h1>Báo cáo Migration Dry-Run</h1>
    <div class="disclaimer">
        <strong>Lưu ý:</strong> Báo cáo này là bản xem trước trên text ghép nối. Khi migrate thật, biên các thẻ XML có thể gây lệch đôi chút (dùng cờ <code>--verbose</code> để debug nếu nghi ngờ).
    </div>
    
    <div class="summary-box">
        <div class="card">
            <h3>Tổng số file</h3>
            <p>{total_files}</p>
        </div>
        <div class="card">
            <h3>File có thay đổi</h3>
            <p>{modified_files}</p>
        </div>
        <div class="card">
            <h3>Tổng số thay đổi</h3>
            <p>{total_changes}</p>
        </div>
    </div>
    
    <table>
        <thead>
            <tr>
                <th style="width: 25%;">File</th>
                <th style="width: 35%;">Đoạn văn sau thay thế (Preview)</th>
                <th style="width: 40%;">Chi tiết biến được thay</th>
            </tr>
        </thead>
        <tbody>
    """
    
    for filepath, res in run_logs.items():
        basename = os.path.basename(filepath)
        if not res['modified']:
            html_content += f"""
            <tr>
                <td><strong>{basename}</strong></td>
                <td colspan="2" style="color: #95a5a6; font-style: italic;">- Không có thay đổi -</td>
            </tr>
            """
            continue
            
        # Thống kê cảnh báo
        warn_html = ""
        if res['warnings']:
            warn_html = "<div class='warning-box'>" + "<br>".join(res['warnings']) + "</div>"
            
        rows_span = len(res['logs'])
        for idx, log in enumerate(res['logs']):
            td_file = f'<td rowspan="{rows_span}"><strong>{basename}</strong>{warn_html}</td>' if idx == 0 else ''
            
            detail_items = []
            for m in log['matches']:
                detail_items.append(f"<span class='del'>{m['original']}</span> &rarr; <span class='ins'>{m['placeholder']}</span>")
            details = "<br>".join(detail_items)
            
            html_content += f"""
            <tr>
                {td_file}
                <td>{log['original_p']}</td>
                <td>{details}</td>
            </tr>
            """
            
    html_content += """
        </tbody>
    </table>
</body>
</html>
    """
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"HTML Report generated at: {report_path}")

def generate_excel_report(report_dir, run_logs, timestamp):
    """Sinh báo cáo Excel để phục vụ audit."""
    os.makedirs(report_dir, exist_ok=True)
    report_path = os.path.join(report_dir, f"dryrun_{timestamp}.xlsx")
    
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Migration Audit"
    
    # Headers
    headers = ["File", "Đoạn văn (Preview)", "Nguyên bản mẫu", "Placeholder thay thế", "Tổng thay đổi"]
    ws.append(headers)
    
    # Style
    header_fill = PatternFill(start_color="F1F2F6", end_color="F1F2F6", fill_type="solid")
    font_bold = Font(name="Segoe UI", size=11, bold=True)
    font_regular = Font(name="Segoe UI", size=10)
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = font_bold
        
    red_fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")
    green_fill = PatternFill(start_color="E8FFE8", end_color="E8FFE8", fill_type="solid")
    
    row_count = 2
    for filepath, res in run_logs.items():
        basename = os.path.basename(filepath)
        if not res['modified']:
            ws.append([basename, "- Không có thay đổi -", "", "", 0])
            ws.cell(row=row_count, column=1).font = font_regular
            ws.cell(row=row_count, column=2).font = font_regular
            row_count += 1
            continue
            
        for log in res['logs']:
            for m in log['matches']:
                ws.append([
                    basename,
                    log['original_p'],
                    m['original'],
                    m['placeholder'],
                    len(log['matches'])
                ])
                # Apply styles
                for c in range(1, 6):
                    ws.cell(row=row_count, column=c).font = font_regular
                ws.cell(row=row_count, column=2).fill = red_fill
                ws.cell(row=row_count, column=4).fill = green_fill
                row_count += 1
                
    # Auto-fit columns và auto filter
    ws.auto_filter.ref = f"A1:E{row_count-1}"
    ws.freeze_panes = "A2"
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)
        
    wb.save(report_path)
    print(f"Excel Report generated at: {report_path}")

def main():
    parser = argparse.ArgumentParser(description="Chuyển đổi Text thô sang Jinja2 Placeholder trong file Word (.docx)")
    parser.add_argument("--excel", required=True, help="Đường dẫn file Excel chứa data mẫu")
    parser.add_argument("--row", type=int, required=True, help="Index hàng dữ liệu (1-based, bỏ qua header)")
    parser.add_argument("--docx-dir", required=True, help="Thư mục chứa các file docx cần chuyển đổi")
    parser.add_argument("--sheet", help="Tên sheet chứa data mẫu")
    parser.add_argument("--config-sheet", help="Tên sheet Config của KisorDoc")
    parser.add_argument("--case-insensitive", action="store_true", help="Bật so khớp không phân biệt hoa/thường")
    parser.add_argument("--min-length", type=int, default=3, help="Độ dài tối thiểu của text mẫu để xử lý")
    parser.add_argument("--dense-threshold", type=int, default=15, help="Ngưỡng cảnh báo lặp lại quá nhiều lần")
    parser.add_argument("--include", help="Glob pattern file cần bao gồm (VD: *.docx)")
    parser.add_argument("--exclude", help="Glob pattern file cần bỏ qua")
    parser.add_argument("--max-files", type=int, help="Giới hạn số file xử lý tối đa")
    parser.add_argument("--verbose", action="store_true", help="Hiển thị log chi tiết")
    parser.add_argument("--dry-run", action="store_true", help="Chạy nháp không lưu file Word")
    parser.add_argument("--report-dir", help="Thư mục xuất báo cáo HTML/Excel")
    
    args = parser.parse_args()
    
    if not os.path.exists(args.excel):
        print(f"[Error] Excel file not found: {args.excel}")
        sys.exit(1)
        
    docx_dir = args.docx_dir
    
    if not os.path.isdir(docx_dir):
        print(f"[Error] Template directory not found: {docx_dir}")
        sys.exit(1)
        
    if args.verbose:
        print("Đang đọc Excel và xây dựng mapping...")
        
    # 1. Load mapping
    try:
        sorted_mapping, collisions = load_mapping(
            excel_path=args.excel,
            row_idx=args.row,
            sheet_name=args.sheet,
            config_sheet_name=args.config_sheet,
            min_length=args.min_length,
            verbose=args.verbose
        )
    except Exception as e:
        print(f"[Error] Failed to parse Excel: {e}")
        sys.exit(1)
        
    if args.verbose:
        print(f"Tổng số cặp mapping hợp lệ: {len(sorted_mapping)}")
        for sample, placeholder in sorted_mapping:
            print(f"  '{sample}' -> '{placeholder}'")
            
    # 2. Duyệt tìm các file docx
    all_docx = []
    include_pat = re.compile(args.include.replace('*', '.*')) if args.include else None
    exclude_pat = re.compile(args.exclude.replace('*', '.*')) if args.exclude else None
    
    for root, dirs, files in os.walk(docx_dir):
        for f in files:
            if not f.endswith('.docx') or f.endswith('.bak.docx') or f.startswith('~$'):
                continue
            # Lọc include/exclude
            if include_pat and not include_pat.match(f):
                continue
            if exclude_pat and exclude_pat.match(f):
                continue
            all_docx.append(os.path.join(root, f))
            
    if args.max_files:
        all_docx = all_docx[:args.max_files]
        
    if not all_docx:
        print("Không tìm thấy file .docx nào thỏa mãn bộ lọc.")
        sys.exit(0)
        
    # 3. Tiến hành xử lý từng file
    run_logs = {}
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    print(f"Bắt đầu xử lý {len(all_docx)} file docx...")
    for docx_path in all_docx:
        if args.verbose:
            print(f"Đang xử lý: {docx_path}")
        try:
            modified, file_logs, warnings = process_docx(
                docx_path=docx_path,
                sorted_mapping=sorted_mapping,
                case_insensitive=args.case_insensitive,
                dense_threshold=args.dense_threshold,
                dry_run=args.dry_run,
                verbose=args.verbose
            )
            run_logs[docx_path] = {
                'modified': modified,
                'logs': file_logs,
                'warnings': warnings
            }
            if modified:
                print(f"  -> {'[Dry-run]' if args.dry_run else '[Thành công]'} Thay đổi được phát hiện trong file: {os.path.basename(docx_path)}")
        except Exception as e:
            print(f"[Error] Failed to process {docx_path}: {e}")
            if not args.dry_run:
                # Đảm bảo lỗi không làm mất file
                sys.exit(1)
                
    # 4. Xuất báo cáo
    report_dir = args.report_dir
    if report_dir or args.dry_run:
        r_dir = report_dir or "reports"
        try:
            generate_html_report(r_dir, run_logs, timestamp)
            generate_excel_report(r_dir, run_logs, timestamp)
        except Exception as e:
            print(f"[Warning] Failed to generate reports: {e}")
            
    print("Hoàn thành quá trình xử lý.")

if __name__ == "__main__":
    main()
