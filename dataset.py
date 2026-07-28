import duckdb
import pandas as pd
import openpyxl
from pathlib import Path
from config import AppConfig


class DataSet:
    def __init__(self, config: AppConfig):
        self.config = config
        self.conn = duckdb.connect(":memory:")
        self.table_names: list[str] = []
        self._load()

    def _load(self):
        exception_prefix = self.config.ExceptionSheet
        for xlsx_file in sorted(self.config.data_path.glob("*.xlsx")):
            if xlsx_file.name.startswith("~$"):
                continue
            try:
                wb = openpyxl.load_workbook(xlsx_file, read_only=True)
            except Exception:
                continue
            for sheet_name in wb.sheetnames:
                stripped = sheet_name.strip()
                if stripped.startswith(exception_prefix):
                    continue
                if stripped in self.table_names:
                    continue
                df = _sheet_to_dataframe(xlsx_file, stripped)
                if df is not None and not df.empty:
                    safe_name = _safe_table_name(stripped)
                    self.conn.register(safe_name, df)
                    self.table_names.append(safe_name)
            wb.close()

    def query(self, sql: str) -> list[dict]:
        return self.conn.execute(sql).fetchdf().to_dict(orient="records")

    def get_table(self, name: str) -> pd.DataFrame | None:
        safe = _safe_table_name(name)
        if safe in self.table_names:
            return self.conn.execute(f"SELECT * FROM \"{safe}\"").fetchdf()
        return None

    def close(self):
        self.conn.close()


def _sheet_to_dataframe(xlsx_path: Path, sheet_name: str) -> pd.DataFrame | None:
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            for sn in wb.sheetnames:
                if sn.strip() == sheet_name:
                    ws = wb[sn]
                    break
            else:
                wb.close()
                return None
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return None
        header = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(rows[0])]
        data_rows = []
        for row in rows[1:]:
            if any(v is not None for v in row):
                data_rows.append(row)
        if not data_rows:
            return None
        df = pd.DataFrame(data_rows, columns=header)
        df = df.dropna(how="all")
        return df
    except Exception:
        return None


def _safe_table_name(name: str) -> str:
    safe = "".join(c if c.isalnum() or c == "_" else "_" for c in name)
    if not safe or safe[0].isdigit():
        safe = "t_" + safe
    return safe
