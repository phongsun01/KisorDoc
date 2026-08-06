import re
import math
from pathlib import Path
import pandas as pd
import openpyxl
from .config import AppConfig
from .dataset import DataSet
from .utils import (
    _str,
    clean_config_key,
    safe_format,
    resolve_sheet_query,
    _parse_repeat_sheet_config,
    _parse_repeat_key_id,
    _parse_price,
    _parse_row_range,
    _safe_eval_condition,
    validate_sql_identifier,
)


class KisorService:
    def __init__(self, config: AppConfig, ds: DataSet):
        self.config = config
        self.ds = ds

    def get_options(self) -> list[str]:
        rows = self.ds.query("SELECT Key, Value FROM Options ORDER BY Key")
        return [f"{_str(r['Key'])}: {_str(r['Value'])}" for r in rows]

    def get_option_config(self, option_key: str) -> dict:
        if not option_key:
            return {}
        opt_code = option_key.split(":")[0].strip() if ":" in option_key else option_key.strip()
        try:
            rows = self.ds.query("SELECT * FROM Options")
        except Exception:
            rows = []
        for r in rows:
            if _str(r.get("Key")) == opt_code:
                return {
                    "sheet": _str(r.get("Sheet"), self.config.DataSheet),
                    "show": _str(r.get("Show"), self.config.DefaultShow),
                    "key_id": _str(r.get("KeyId"), self.config.DefaultKeyId),
                    "config_range": _str(r.get("Config"), ""),
                    "type": _str(r.get("Type"), ""),
                    "sort_col": _str(r.get("SortCol"), ""),
                }
        return {
            "sheet": self.config.DataSheet,
            "show": self.config.DefaultShow,
            "key_id": self.config.DefaultKeyId,
            "config_range": "",
            "type": "",
            "sort_col": "",
        }

    def get_config_for_option(self, option_key: str) -> list[dict]:
        opt_config = self.get_option_config(option_key)
        cfg_range = opt_config.get("config_range", "")
        parsed = _parse_row_range(cfg_range)
        if parsed:
            start, end = parsed
            return self.ds.query_rows("Config", start, end)
        try:
            return self.ds.query("SELECT * FROM Config")
        except Exception:
            return []

    def get_all_option_templates(self, option_key: str) -> list[str]:
        if not option_key:
            return []
        opt = option_key.split(":")[0].strip() if ":" in option_key else option_key.strip()
        try:
            ws_rows = self.ds.query("SELECT * FROM Workflow")
        except Exception:
            ws_rows = []
        return [str(r.get("Name", "")) for r in ws_rows if _str(r.get("Option")) == opt]

    def check_condition(self, condition_str: str, raw_row: dict, config_mappings: list[dict]) -> bool:
        if not condition_str or condition_str.strip() == "" or condition_str.upper() == "ALL":
            return True
        
        placeholders = re.findall(r"\{(.*?)\}", condition_str)
        
        key_to_col = {}
        for mapping in config_mappings:
            k = clean_config_key(_str(mapping.get("Key")))
            c = _str(mapping.get("Value"))
            if k and c:
                key_to_col[k] = c

        eval_context = {}
        patched_condition = condition_str
        
        for idx, p in enumerate(placeholders):
            p_clean = p.strip()
            safe_var = f"var_{idx}"
            
            col_name = key_to_col.get(p_clean, p_clean)
            val = raw_row.get(col_name)
            
            parsed_val = None
            if val is not None:
                try:
                    is_na = pd.isna(val)
                except (TypeError, ValueError):
                    is_na = False

                if not is_na:
                    parsed_val = _parse_price(val)
                    if parsed_val is None:
                        s_val = str(val).strip()
                        if s_val:
                            parsed_val = s_val
            
            eval_context[safe_var] = parsed_val
            patched_condition = patched_condition.replace(f"{{{p}}}", safe_var)

        try:
            return _safe_eval_condition(patched_condition, eval_context)
        except Exception as e:
            print(f"⚠️  Lỗi cú pháp điều kiện lọc '{condition_str}' (biểu thức sau chuyển đổi: '{patched_condition}'): {e}")
            return False

    def get_packages(self, option_key: str) -> list[str]:
        if not option_key:
            return []
        opt_config = self.get_option_config(option_key)
        sheet = opt_config.get("sheet", self.config.DataSheet)
        show_format = opt_config.get("show", "")
        if "|" in show_format:
            show_format = show_format.split("|")[0].strip()
        
        if opt_config.get("type") == "Repeat":
            ls, _, _ = _parse_repeat_sheet_config(opt_config)
            sql = f'SELECT * FROM "{ls}"' if ls else resolve_sheet_query(sheet)
        else:
            sql = resolve_sheet_query(sheet)
        sort_col = opt_config.get("sort_col", "")
        try:
            if sort_col:
                rows = self.ds.query(f"SELECT * FROM ({sql}) ORDER BY CAST(\"{sort_col}\" AS INTEGER)")
            else:
                rows = self.ds.query(sql)
        except Exception:
            try:
                rows = self.ds.query(sql)
            except Exception:
                rows = []
            
        packages = []
        for r in rows:
            label = safe_format(show_format, r)
            if label:
                packages.append(label)
        return packages

    def get_package_details(self, option_key: str, package_label: str, sheet_rows: list[dict] | None = None) -> dict:
        if not option_key or not package_label:
            return {}
        opt_config = self.get_option_config(option_key)
        show_format = opt_config.get("show", "")
        if "|" in show_format:
            show_format = show_format.split("|")[0].strip()

        if sheet_rows is not None:
            rows = sheet_rows
        else:
            sheet = opt_config.get("sheet", self.config.DataSheet)
            sql = resolve_sheet_query(sheet)
            try:
                rows = self.ds.query(sql)
            except Exception:
                rows = []
            
        for r in rows:
            label = safe_format(show_format, r)
            if label == package_label:
                details = {}
                for col_name, val in r.items():
                    if val is not None and str(val).strip() != "":
                        details[col_name] = _str(val)
                return details
        return {}

    def get_package_excel_file(self, goi_thau_id: str, key_id: str | None = None) -> Path | None:
        if key_id is None:
            key_id = self.config.DefaultKeyId
        key_id = validate_sql_identifier(key_id)
        try:
            rows = self.ds.query(f'SELECT DISTINCT File FROM Tables WHERE "{key_id}" = ? AND File IS NOT NULL AND File != \'nan\'', (goi_thau_id,))
            if rows:
                filename = str(rows[0]['File']).strip()
                if filename:
                    xlsx_files = sorted(self.config.data_path.glob("*.xlsx"))
                    for f in xlsx_files:
                        if filename.lower() in f.name.lower():
                            return f
                    # Fallback to DanhMucFile
                    danh_muc_file = next(
                        (f for f in xlsx_files if self.config.DanhMucFile.lower() in f.stem.lower()),
                        xlsx_files[0] if xlsx_files else None
                    )
                    return danh_muc_file
        except Exception as e:
            print(f"⚠️ Lỗi tìm file Excel cho gói thầu {goi_thau_id}: {e}")
        return None

    def get_workflow_templates(self, option_key: str, package_label: str, sheet_rows: list[dict] | None = None) -> list[dict]:
        if not option_key or not package_label or package_label.strip() == "":
            return []
        opt = option_key.split(":")[0].strip() if ":" in option_key else option_key.strip()
        
        try:
            ws_rows = self.ds.query("SELECT * FROM Workflow")
        except Exception:
            ws_rows = []
        wf_rows = [r for r in ws_rows if _str(r.get("Option")) == opt]
        if not wf_rows:
            return []
            
        opt_config = self.get_option_config(option_key)
        key_id = opt_config.get("key_id") or self.config.DefaultKeyId
        show_format = opt_config.get("show", "")
        if "|" in show_format:
            show_format = show_format.split("|")[0].strip()
        
        if sheet_rows is not None:
            main_rows = sheet_rows
        else:
            sheet = opt_config.get("sheet", self.config.DataSheet)
            if opt_config.get("type") == "Repeat":
                left_sheet, _, _ = _parse_repeat_sheet_config(opt_config)
                sql = f'SELECT * FROM "{left_sheet}"' if left_sheet else resolve_sheet_query(sheet)
            else:
                sql = resolve_sheet_query(sheet)
            try:
                main_rows = self.ds.query(sql)
            except Exception:
                main_rows = []
            
        selected_pkg = None
        for r in main_rows:
            label = safe_format(show_format, r)
            if label == package_label:
                selected_pkg = r
                break
                
        if not selected_pkg:
            return []
            
        config_mappings = self.get_config_for_option(option_key)
            
        filtered = []
        for r in wf_rows:
            condition_str = _str(r.get("Condition", ""))
            if not condition_str or condition_str.upper() == "ALL":
                filtered.append(r)
            else:
                if self.check_condition(condition_str, selected_pkg, config_mappings):
                    filtered.append(r)
        return filtered

    def get_repeat_members(self, goi_thau_id: str, group_type: str, option_key: str = None) -> list[str]:
        member_show_format = ""
        left_sheet = right_sheet = join_key = ""
        if option_key:
            opt_config = self.get_option_config(option_key)
            show_format = opt_config.get("show", "")
            if "|" in show_format:
                member_show_format = show_format.split("|", 1)[1].strip()
            left_sheet, right_sheet, join_key = _parse_repeat_sheet_config(opt_config)

        if not right_sheet:
            print(f"⚠️ get_repeat_members: không có right_sheet trong option '{option_key}'")
            return []

        # FIX SVC-01: tên bảng _Goc phải qua _safe_table_name để khớp với tên
        # mà DataSet._load() đã register (sheet có space/ký tự đặc biệt sẽ bị đổi)
        from .dataset import _safe_table_name as _stn
        goc_table = _stn(right_sheet + "_Goc")

        try:
            if join_key and goi_thau_id:
                try:
                    rows = self.ds.query(f'SELECT * FROM "{goc_table}" WHERE "{join_key}" = ?', (goi_thau_id,))
                except Exception:
                    rows = self.ds.query(f'SELECT * FROM "{goc_table}"')
            else:
                rows = self.ds.query(f'SELECT * FROM "{goc_table}"')
        except Exception as e:
            print(f"❌ get_repeat_members query error: {e}")
            return []

        members = []
        for r in rows:
            if member_show_format:
                label = safe_format(member_show_format, {k: _str(v) for k, v in r.items()})
                if label.strip():
                    members.append(label.strip())
            else:
                vals = list(r.values())
                name = _str(vals[1]) if len(vals) > 1 else ""
                if name:
                    members.append(name)
        return members

    def register_temporary_tcgttd(self, goi_thau_id: str, selected_member_names: list[str],
                                   group_name: str, key_id: str, option_key: str = None) -> bool:
        opt_config = self.get_option_config(option_key) if option_key else {}
        left_sheet, right_sheet, join_key = _parse_repeat_sheet_config(opt_config)
        left_key, right_key = _parse_repeat_key_id(key_id)

        if not right_sheet:
            print(f"⚠️ register_temporary_tcgttd: không có right_sheet trong option '{option_key}'")
            return False

        member_show_format = ""
        show_format = opt_config.get("show", "")
        if "|" in show_format:
            member_show_format = show_format.split("|", 1)[1].strip()

        # FIX SVC-01: tên bảng _Goc phải qua _safe_table_name (nhất quán với DataSet._load)
        from .dataset import _safe_table_name as _stn
        goc_table = _stn(right_sheet + "_Goc")

        try:
            if join_key and goi_thau_id:
                try:
                    all_rows = self.ds.query(f'SELECT * FROM "{goc_table}" WHERE "{join_key}" = ?', (goi_thau_id,))
                except Exception:
                    all_rows = self.ds.query(f'SELECT * FROM "{goc_table}"')
            else:
                all_rows = self.ds.query(f'SELECT * FROM "{goc_table}"')
        except Exception as e:
            print(f"❌ register_temporary_tcgttd query error: {e}")
            return False

        matched = []
        for r in all_rows:
            if member_show_format:
                label = safe_format(member_show_format, {k: _str(v) for k, v in r.items()})
            else:
                vals = list(r.values())
                label = _str(vals[1]) if len(vals) > 1 else ""
            if label.strip() in selected_member_names:
                matched.append(dict(r))

        if not matched:
            print(f"⚠️ register_temporary_tcgttd: không tìm thấy {selected_member_names}")
            return False

        df_temp = pd.DataFrame(matched)
        if join_key and join_key not in df_temp.columns:
            df_temp[join_key] = goi_thau_id
        self.ds.conn.register(right_sheet, df_temp)
        return True

    def run_preview(self, option_key: str, package_label: str, selected_templates: list[str]) -> str:
        if not option_key or not package_label or not selected_templates:
            return "⚠️ Chọn đủ quy trình, gói thầu và ít nhất 1 template trước"

        opt_config = self.get_option_config(option_key)
        sheet = opt_config.get("sheet", self.config.DataSheet)
        key_id = opt_config.get("key_id") or self.config.DefaultKeyId
        show_format = opt_config.get("show", "")
        if "|" in show_format:
            show_format = show_format.split("|")[0].strip()

        sql = resolve_sheet_query(sheet)
        goi_thau_rows = self.ds.query(sql)
        selected_pkg = next((
            r for r in goi_thau_rows
            if safe_format(show_format, r) == package_label
        ), None)
        if not selected_pkg:
            return "❌ Không tìm thấy dòng dữ liệu tương ứng"

        config_rows = self.get_config_for_option(option_key)
        context_keys = set()
        missing_keys = []

        for r in config_rows:
            key = _str(r.get("Key"))
            col = _str(r.get("Value"))
            if not key or not col:
                continue
            clean_key = clean_config_key(key)

            raw_value = selected_pkg.get(col)
            try:
                is_na = pd.isna(raw_value)
            except (TypeError, ValueError):
                is_na = False

            context_keys.add(clean_key)
            if is_na or raw_value is None or str(raw_value).strip() == "":
                missing_keys.append(clean_key)

        lines = []
        if missing_keys:
            lines.append(
                f"✅ Context: {len(context_keys)} key  |  "
                f"⚠️ Thiếu data: {', '.join(missing_keys)}"
            )
        else:
            lines.append(f"✅ Context: {len(context_keys)} key – đầy đủ")

        try:
            tables_rows = self.ds.query("SELECT * FROM Tables")
        except Exception:
            tables_rows = []

        # Fix composite key_id issue
        left_key, _ = _parse_repeat_key_id(key_id)
        goi_thau_id = _str(selected_pkg.get(left_key))

        xlsx_files = sorted(self.config.data_path.glob("*.xlsx"))
        danh_muc_file = next(
            (f for f in xlsx_files if self.config.DanhMucFile.lower() in f.stem.lower()),
            xlsx_files[0] if xlsx_files else None
        )

        wb = None
        if danh_muc_file and danh_muc_file.exists():
            try:
                wb = openpyxl.load_workbook(danh_muc_file, read_only=True, data_only=True)
            except Exception:
                wb = None

        table_lines = []
        for t in tables_rows:
            t_id = t.get(left_key)  # also using left_key dynamically here
            if _str(t_id) != goi_thau_id:
                continue
            name = _str(t.get("Name", ""))
            tbl_sheet = _str(t.get("Sheet", ""))
            range_ = _str(t.get("Range", ""))
            hide = _str(t.get("Hide", ""))

            row_count = "?"
            if wb and tbl_sheet and tbl_sheet in wb.sheetnames:
                try:
                    ws = wb[tbl_sheet]
                    if ":" in range_:
                        parts = range_.split(":")
                        end = parts[1]
                        digits = "".join(c for c in end if c.isdigit())
                        row_count = digits if digits else str(ws.max_row)
                    else:
                        row_count = str(ws.max_row)
                except Exception:
                    pass

            hide_str = f"  |  ẩn: {hide}" if hide else ""
            table_lines.append(
                f"📋 {name} → {tbl_sheet}  {range_}  {row_count} dòng{hide_str}"
            )

        if wb:
            try:
                wb.close()
            except Exception:
                pass

        if table_lines:
            lines.extend(table_lines)
        elif tables_rows:
            lines.append("📋 Tables: không có dòng nào khớp với gói thầu này")
        else:
            lines.append("📋 Tables: (không có sheet Tables)")

        return "\n".join(lines)
