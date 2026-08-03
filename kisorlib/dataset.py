"""
PATCH: dataset.py
BUG FIX: Mở file Excel 2 lần mỗi sheet → gộp lại thành 1 lần đọc trong _load()
"""

import duckdb
import pandas as pd
import openpyxl
from pathlib import Path
from .config import AppConfig


class DataSet:
    def __init__(self, config: AppConfig):
        self.config = config
        self.conn = duckdb.connect(":memory:")
        self.table_names: list[str] = []
        self._load()

    def _load(self):
        exception_prefix = self.config.ExceptionSheet
        for xlsx_file in sorted(self.config.data_path.glob("*.xlsx")):
            if xlsx_file.name.startswith("~$") or xlsx_file.name.startswith(exception_prefix):
                continue
            # FIX: Chỉ mở file 1 lần duy nhất, dùng data_only=True luôn
            try:
                wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
            except Exception:
                continue

            for sheet_name in wb.sheetnames:
                stripped = sheet_name.strip()
                if stripped.startswith(exception_prefix):
                    continue
                safe_name = _safe_table_name(stripped)

                df = _ws_to_dataframe(wb[sheet_name])
                if df is not None and not df.empty:
                    if safe_name in self.table_names:
                        try:
                            existing_df = self.conn.execute(f'SELECT * FROM "{safe_name}"').fetchdf()
                            combined_df = pd.concat([existing_df, df], ignore_index=True)
                            self.conn.register(safe_name, combined_df)
                        except Exception as merge_err:
                            print(f"⚠️  Lỗi gộp sheet trùng tên '{safe_name}': {merge_err}")
                    else:
                        self.conn.register(safe_name, df)
                        self.table_names.append(safe_name)

            wb.close()

    def query(self, sql: str) -> list[dict]:
        df = self.conn.execute(sql).fetchdf()
        cols = list(df.columns)
        if len(cols) != len(set(cols)):
            dupes = {c for c in cols if cols.count(c) > 1}
            print(f"⚠️  Cảnh báo: Phát hiện trùng tên cột {dupes} khi thực thi liên kết (Join). Vui lòng điều chỉnh lại tên cột trong Excel để tránh ghi đè dữ liệu.")
        return df.to_dict(orient="records")

    def query_rows(self, sheet_name: str, row_start: int, row_end: int) -> list[dict]:
        cache_key = (sheet_name, row_start, row_end)
        if not hasattr(self, "_rows_cache"):
            self._rows_cache = {}
        if cache_key in self._rows_cache:
            return self._rows_cache[cache_key]
        res = self._query_rows_impl(sheet_name, row_start, row_end)
        self._rows_cache[cache_key] = res
        return res

    def _query_rows_impl(self, sheet_name: str, row_start: int, row_end: int) -> list[dict]:
        xlsx_files = sorted(self.config.data_path.glob("*.xlsx"))
        actual_file = None
        clean_sheet = sheet_name.strip()
        for f in xlsx_files:
            if f.name.startswith("~$"):
                continue
            try:
                wb = openpyxl.load_workbook(f, read_only=True)
                if any(s.strip() == clean_sheet for s in wb.sheetnames):
                    actual_file = f
                    wb.close()
                    break
                wb.close()
            except Exception:
                continue

        if not actual_file:
            print(f"⚠️  Không tìm thấy sheet '{sheet_name}' trong bất kỳ file Excel nào")
            return []

        try:
            wb = openpyxl.load_workbook(actual_file, read_only=True, data_only=True)
            actual_sheet_name = next(s for s in wb.sheetnames if s.strip() == clean_sheet)
            ws = wb[actual_sheet_name]

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                wb.close()
                return []
            headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(header_row)]

            data_rows = []
            for r in ws.iter_rows(min_row=row_start, max_row=row_end, values_only=True):
                if any(v is not None for v in r):
                    row_dict = {}
                    for i, val in enumerate(r):
                        if i < len(headers):
                            row_dict[headers[i]] = val
                    data_rows.append(row_dict)
            wb.close()
            return data_rows
        except Exception as e:
            print(f"❌ Lỗi query_rows cho sheet '{sheet_name}': {e}")
            return []

    def get_table(self, name: str) -> pd.DataFrame | None:
        safe = _safe_table_name(name)
        if safe in self.table_names:
            return self.conn.execute(f'SELECT * FROM "{safe}"').fetchdf()
        return None

    def close(self):
        self.conn.close()


def _ws_to_dataframe(ws) -> pd.DataFrame | None:
    """
    FIX: Nhận worksheet object trực tiếp thay vì mở lại file.
    Dùng chung cho cả _load() — không cần _sheet_to_dataframe() riêng nữa.
    """
    try:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None
        header = [
            str(c) if c is not None else f"Col{i}"
            for i, c in enumerate(rows[0])
        ]
        data_rows = [row for row in rows[1:] if any(v is not None for v in row)]
        if not data_rows:
            return None
        df = pd.DataFrame(data_rows, columns=header)
        df = df.dropna(how="all")
        return df
    except Exception:
        return None


def _safe_table_name(name: str) -> str:
    safe = "".join(c if c.isalnum() or c == "_" else "_" for c in name)
    if not safe or safe[0].isdigit():
        safe = "t_" + safe
    return safe
