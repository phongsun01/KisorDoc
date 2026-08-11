"""
kisorlib/text_migrator.py
─────────────────────────
Core logic migrate text thuần trong .docx → Jinja2 placeholder {{ TenBien }}.

Bài toán ngược với migrator.py:
  - migrator.py  : <<TenBien>>  → {{ TenBien }}   (placeholder → Jinja2)
  - text_migrator: "Nguyễn Văn A" → {{ HoTen }}   (giá trị mẫu → Jinja2)

Mapping lấy từ:
  1. Config sheet KisorDoc (Key = tên placeholder, Value = tên cột Excel)
  2. Fallback: slug bỏ dấu từ header Excel (unicodedata)

Hỗ trợ:
  - dry_run=True      : chỉ phân tích, không ghi file
  - on_progress       : callback(dict) cho UI streaming (cùng interface với migrator.py)
  - case_insensitive  : khớp không phân biệt hoa/thường
  - dense_threshold   : cảnh báo nếu 1 giá trị xuất hiện quá nhiều lần

Không import Gradio.
"""

from __future__ import annotations

import re
import os
import shutil
import zipfile
import unicodedata
import datetime
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

try:
    from lxml import etree
except ImportError as e:
    raise ImportError("Cần cài lxml: pip install lxml") from e

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise ImportError("Cần cài openpyxl: pip install openpyxl") from e

# Import clean_config_key từ kisorlib.utils (cùng package)
try:
    from kisorlib.utils import clean_config_key
except ImportError:
    def clean_config_key(key: str) -> str:
        """Fallback minimal khi chạy ngoài môi trường kisorlib."""
        clean = key.strip("<>{}| ")
        for suffix in (".Date.Long", ".Date.long", ".date_long"):
            if clean.endswith(suffix):
                return clean[: -len(suffix)] + "_Date"
        if clean.lower().endswith(".date"):
            return clean[:-5] + "_Date"
        for s in (".Day", ".day", ".Month", ".month", ".Year", ".year"):
            if clean.endswith(s):
                return clean[: -len(s)] + "_Date"
        for s in (".Upper", ".upper", ".Number", ".number"):
            if clean.endswith(s):
                clean = clean[: -len(s)]
                break
        if "|" in clean:
            clean = clean.split("|")[0].strip()
        return clean


# ──────────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────────

W_NS  = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
W     = f"{{{W_NS}}}"
NAMESPACES = {"w": W_NS}

# Các file XML bên trong .docx cần xử lý
_TARGET_XML_PATTERNS = ("document", "header", "footer", "footnote", "endnote")

# Regex nhận diện placeholder Jinja2 đã có (dùng cho idempotency check)
_JINJA2_RE = re.compile(r"\{\{.*?\}\}", re.DOTALL)


# ──────────────────────────────────────────────────────────────────────────────
# Result types  (tương thích interface với migrator.py)
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class TextChange:
    """Một lần thay thế text mẫu → placeholder."""
    original:    str           # Giá trị mẫu bị thay
    placeholder: str           # Placeholder Jinja2 thay vào
    paragraph:   str           # Text đoạn văn sau khi thay (để preview)
    location:    str = ""      # Tên file XML bên trong docx


@dataclass
class TextFileResult:
    """Kết quả xử lý một file .docx — cùng shape với FileResult của migrator.py."""
    path:            Path
    success:         bool
    changed:         bool                  = False
    changes:         list[TextChange]      = field(default_factory=list)
    warnings:        list[str]             = field(default_factory=list)
    error:           Optional[str]         = None
    backed_up_to:    Optional[Path]        = None


OnProgress = Optional[Callable[[dict], None]]


def _emit(cb: OnProgress, level: str, message: str, **extra) -> None:
    if cb is None:
        return
    try:
        cb({"level": level, "message": message, **extra})
    except Exception:
        pass


# ──────────────────────────────────────────────────────────────────────────────
# Slug & mapping helpers
# ──────────────────────────────────────────────────────────────────────────────

def to_slug(text: str) -> str:
    """
    'Họ và Tên'  →  'Ho_va_Ten'
    Giữ nguyên hoa/thường, thay ký tự đặc biệt/khoảng trắng bằng '_'.
    """
    if not text:
        return ""
    s = str(text).replace("đ", "d").replace("Đ", "D")
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def load_mapping(
    excel_path:        str | Path,
    row_idx:           int,
    sheet_name:        Optional[str] = None,
    config_sheet_name: Optional[str] = None,
    min_length:        int = 3,
    verbose:           bool = False,
    on_progress:       OnProgress = None,
) -> tuple[list[tuple[str, str]], list[tuple[str, str, str]]]:
    """
    Đọc Excel → build mapping { giá trị mẫu → "{{ placeholder }}" }.

    Returns:
        sorted_mapping : list[(value, placeholder)] sắp xếp theo len(value) giảm dần
        collisions     : list[(value, first_placeholder, skipped_placeholder)]
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]

    headers   = [cell.value for cell in ws[1]]
    target_row = ws[row_idx + 1]   # row_idx là 1-based không tính header
    row_values = [cell.value for cell in target_row]

    # ── Đọc Config sheet ──────────────────────────────────────────────────────
    config_map: dict[str, str] = {}          # norm(col_name) → Key
    cfg_sheet  = config_sheet_name or "Config"
    if cfg_sheet in wb.sheetnames:
        c_ws = wb[cfg_sheet]
        c_hdrs = [cell.value for cell in c_ws[1]]
        try:
            ki = c_hdrs.index("Key")
            vi = c_hdrs.index("Value")
            for r in range(2, c_ws.max_row + 1):
                k = c_ws.cell(row=r, column=ki + 1).value
                v = c_ws.cell(row=r, column=vi + 1).value
                if k and v:
                    config_map[str(v).strip().lower()] = str(k).strip()
        except ValueError:
            msg = f"Config sheet '{cfg_sheet}' thiếu cột 'Key' hoặc 'Value'."
            _emit(on_progress, "warning", msg)
            if verbose:
                print(f"[Warning] {msg}")
    elif config_sheet_name:
        msg = f"Không tìm thấy sheet Config '{cfg_sheet}'."
        _emit(on_progress, "warning", msg)
        if verbose:
            print(f"[Warning] {msg}")

    # ── Build mapping ─────────────────────────────────────────────────────────
    mapping:    dict[str, str] = {}
    collisions: list[tuple[str, str, str]] = []

    for header, val in zip(headers, row_values):
        if val is None or header is None:
            continue
        val_str = str(val).strip()
        if len(val_str) < min_length:
            continue

        h_norm = str(header).strip().lower()
        if h_norm in config_map:
            raw_key  = config_map[h_norm]
            clean_key = clean_config_key(raw_key)
        else:
            clean_key = to_slug(str(header).strip())

        placeholder = f"{{{{ {clean_key} }}}}"

        if val_str in mapping:
            collisions.append((val_str, mapping[val_str], placeholder))
            msg = f"[Collision] '{val_str}' đã map tới '{mapping[val_str]}', bỏ qua '{placeholder}'."
            _emit(on_progress, "warning", msg)
            if verbose:
                print(msg)
        else:
            mapping[val_str] = placeholder

    # Sort by value length giảm dần — đảm bảo substring không bị thay trước
    sorted_mapping = sorted(mapping.items(), key=lambda x: len(x[0]), reverse=True)
    return sorted_mapping, collisions


# ──────────────────────────────────────────────────────────────────────────────
# XML paragraph helpers
# ──────────────────────────────────────────────────────────────────────────────

def _reconstruct(p_node) -> tuple[str, list[tuple]]:
    """
    Ghép text tất cả <w:t> trong paragraph.
    Trả về (full_text, offset_map).
    offset_map[i] = (r_node, char_offset_in_t, t_node)
    """
    full_text  = ""
    offset_map = []
    for t_node in p_node.xpath(".//w:t", namespaces=NAMESPACES):
        r_node = t_node.getparent()
        t_text = t_node.text or ""
        for i, _ in enumerate(t_text):
            offset_map.append((r_node, i, t_node))
        full_text += t_text
    return full_text, offset_map


def _forbidden_spans(full_text: str) -> list[tuple[int, int]]:
    """Trả về danh sách span của các {{ ... }} đã có — vùng idempotency."""
    return [m.span() for m in _JINJA2_RE.finditer(full_text)]


def _overlaps(span: tuple[int, int], forbidden: list[tuple[int, int]]) -> bool:
    s, e = span
    return any(not (e <= fs or fe <= s) for fs, fe in forbidden)


def _apply_replacement(
    p_node,
    start_t, start_offset: int,
    end_t,   end_offset:   int,
    sample:      str,
    placeholder: str,
) -> None:
    """
    Ghi placeholder vào XML span (start_t, start_offset) → (end_t, end_offset).
    Giữ nguyên formatting, chỉ sửa nội dung text.
    """
    if start_t is end_t:
        # Nằm gọn trong 1 thẻ <w:t>
        t = start_t.text or ""
        start_t.text = t[:start_offset] + placeholder + t[start_offset + len(sample):]
    else:
        # Vắt ngang nhiều run
        st = start_t.text or ""
        start_t.text = st[:start_offset] + placeholder

        all_t = p_node.xpath(".//w:t", namespaces=NAMESPACES)
        try:
            si = all_t.index(start_t)
            ei = all_t.index(end_t)
            for mid in range(si + 1, ei):
                all_t[mid].text = ""
        except ValueError:
            pass

        et = end_t.text or ""
        end_t.text = et[end_offset + 1:]   # trim phần đã dùng, giữ hậu tố


# ──────────────────────────────────────────────────────────────────────────────
# Core paragraph migration
# ──────────────────────────────────────────────────────────────────────────────

def migrate_paragraph(
    p_node,
    sorted_mapping:  list[tuple[str, str]],
    case_insensitive: bool = False,
    dense_threshold:  int  = 15,
    location:        str  = "",
) -> list[TextChange]:
    """
    Xử lý một paragraph XML.
    Trả về list[TextChange] — rỗng nếu không có gì thay đổi.
    """
    full_text, offset_map = _reconstruct(p_node)
    if not full_text or not offset_map:
        return []

    forbidden  = _forbidden_spans(full_text)
    changes: list[TextChange] = []
    match_counts: dict[str, int] = {}

    for sample, placeholder in sorted_mapping:
        flags   = re.IGNORECASE if case_insensitive else 0
        pattern = re.escape(sample)

        for m in reversed(list(re.finditer(pattern, full_text, flags))):
            span = m.span()
            if _overlaps(span, forbidden):
                continue

            si, ei = span
            if si >= len(offset_map) or ei - 1 >= len(offset_map):
                continue

            start_run, start_off, start_t = offset_map[si]
            end_run,   end_off,   end_t   = offset_map[ei - 1]

            _apply_replacement(p_node, start_t, start_off, end_t, end_off, sample, placeholder)

            match_counts[sample] = match_counts.get(sample, 0) + 1
            changes.append(TextChange(
                original=sample,
                placeholder=placeholder,
                paragraph="",   # filled below after reconstruct
                location=location,
            ))

            # Reconstruct sau mỗi thay thế để offset_map và forbidden chính xác
            full_text, offset_map = _reconstruct(p_node)
            forbidden = _forbidden_spans(full_text)

    # Gán paragraph preview cho tất cả changes trong đoạn này
    if changes:
        preview = full_text
        for c in changes:
            c.paragraph = preview

    return changes


# ──────────────────────────────────────────────────────────────────────────────
# XML file-level migration
# ──────────────────────────────────────────────────────────────────────────────

def migrate_xml_bytes(
    data:            bytes,
    sorted_mapping:  list[tuple[str, str]],
    case_insensitive: bool = False,
    dense_threshold:  int  = 15,
    location:        str  = "",
) -> tuple[bytes, list[TextChange], list[str]]:
    """
    Xử lý nội dung bytes của một file XML trong docx.
    Trả về (new_bytes, changes, warnings).
    """
    root    = etree.fromstring(data)
    changes: list[TextChange] = []
    warnings: list[str]       = []
    match_counts: dict[str, int] = {}

    paragraphs = root.xpath("//w:p", namespaces=NAMESPACES)
    for p in paragraphs:
        p_changes = migrate_paragraph(
            p, sorted_mapping,
            case_insensitive=case_insensitive,
            dense_threshold=dense_threshold,
            location=location,
        )
        changes.extend(p_changes)
        for c in p_changes:
            match_counts[c.original] = match_counts.get(c.original, 0) + 1

    # Dense Match Warning
    for sample, count in match_counts.items():
        if count > dense_threshold:
            warnings.append(
                f"[Dense Match] '{sample}' xuất hiện {count} lần trong {location} "
                f"(ngưỡng: {dense_threshold})"
            )

    new_bytes = etree.tostring(root, encoding="utf-8", xml_declaration=True)
    return new_bytes, changes, warnings


# ──────────────────────────────────────────────────────────────────────────────
# File-level migration
# ──────────────────────────────────────────────────────────────────────────────

def _target_xml_files(namelist: list[str]) -> list[str]:
    return [
        f for f in namelist
        if f.startswith("word/") and f.endswith(".xml")
        and any(pat in f for pat in _TARGET_XML_PATTERNS)
    ]


def text_migrate_file(
    path:            Path,
    sorted_mapping:  list[tuple[str, str]],
    *,
    dry_run:          bool       = False,
    backup:           bool       = True,
    case_insensitive: bool       = False,
    dense_threshold:  int        = 15,
    on_progress:      OnProgress = None,
) -> TextFileResult:
    """
    Migrate text mẫu → placeholder trong 1 file .docx.

    Parameters
    ----------
    path            : đường dẫn file .docx
    sorted_mapping  : output của load_mapping()
    dry_run         : nếu True, không ghi file (chỉ phân tích)
    backup          : tạo backup timestamp trước khi ghi đè
    case_insensitive: khớp không phân biệt hoa/thường
    dense_threshold : ngưỡng cảnh báo dense match
    on_progress     : callback(dict) — cùng interface với migrator.py
    """
    result = TextFileResult(path=path, success=False)

    if not path.exists():
        result.error = f"File không tồn tại: {path}"
        _emit(on_progress, "error", result.error, file=path.name)
        return result

    try:
        all_changes:  list[TextChange] = []
        all_warnings: list[str]        = []
        xml_data:     dict[str, bytes] = {}

        # Đọc toàn bộ XML targets
        with zipfile.ZipFile(path, "r") as zin:
            namelist = zin.namelist()
            targets  = _target_xml_files(namelist)
            for t in targets:
                xml_data[t] = zin.read(t)

        # Xử lý từng file XML
        migrated: dict[str, bytes] = {}
        for xml_name, data in xml_data.items():
            new_data, changes, warnings = migrate_xml_bytes(
                data, sorted_mapping,
                case_insensitive=case_insensitive,
                dense_threshold=dense_threshold,
                location=xml_name,
            )
            migrated[xml_name]  = new_data
            all_changes.extend(changes)
            all_warnings.extend(warnings)

        result.changes  = all_changes
        result.warnings = all_warnings
        result.changed  = bool(all_changes)

        if dry_run or not all_changes:
            result.success = True
            label = "[dry-run]" if dry_run else "[no-change]"
            _emit(on_progress, "info",
                  f"{label} {path.name}: {len(all_changes)} thay đổi",
                  file=path.name, count=len(all_changes))
            return result

        # ── Ghi file thật ────────────────────────────────────────────────────
        if backup:
            ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            bak_path = path.with_name(f"{path.stem}.{ts}.bak.docx")
            shutil.copy2(path, bak_path)
            result.backed_up_to = bak_path

        tmp_path = path.with_suffix(".tmp.docx")
        with zipfile.ZipFile(path, "r") as zin:
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename in migrated:
                        zout.writestr(item, migrated[item.filename])
                    else:
                        zout.writestr(item, zin.read(item.filename))

        tmp_path.replace(path)
        result.success = True
        _emit(on_progress, "success",
              f"✓ {path.name}: {len(all_changes)} thay đổi",
              file=path.name, count=len(all_changes))

    except Exception as exc:
        result.error   = f"{type(exc).__name__}: {exc}"
        result.success = False
        _emit(on_progress, "error", f"✗ {path.name}: {result.error}", file=path.name)
        tmp = path.with_suffix(".tmp.docx")
        if tmp.exists():
            try:
                tmp.unlink()
            except Exception:
                pass

    return result


# ──────────────────────────────────────────────────────────────────────────────
# Folder-level migration
# ──────────────────────────────────────────────────────────────────────────────

def text_migrate_folder(
    folder:          Path,
    sorted_mapping:  list[tuple[str, str]],
    *,
    dry_run:          bool       = False,
    backup:           bool       = True,
    recursive:        bool       = True,
    case_insensitive: bool       = False,
    dense_threshold:  int        = 15,
    include_pat:      Optional[re.Pattern] = None,
    exclude_pat:      Optional[re.Pattern] = None,
    max_files:        Optional[int]        = None,
    on_progress:      OnProgress           = None,
) -> list[TextFileResult]:
    """
    Migrate tất cả .docx trong folder.
    Tự động bỏ qua *.bak.docx và file ~$ tạm của Word.
    """
    pattern = "**/*.docx" if recursive else "*.docx"
    files = [
        p for p in sorted(folder.glob(pattern))
        if not p.name.endswith(".bak.docx")
        and not p.name.startswith("~$")
        and "bak" not in p.parts
    ]

    # Lọc include / exclude
    if include_pat:
        files = [p for p in files if include_pat.search(p.name)]
    if exclude_pat:
        files = [p for p in files if not exclude_pat.search(p.name)]
    if max_files:
        files = files[:max_files]

    if not files:
        _emit(on_progress, "warning", f"Không tìm thấy file .docx trong: {folder}")
        return []

    _emit(on_progress, "info",
          f"Tìm thấy {len(files)} file .docx",
          total=len(files))

    results: list[TextFileResult] = []
    for i, f in enumerate(files, 1):
        _emit(on_progress, "info",
              f"[{i}/{len(files)}] {f.name}",
              step=i, total=len(files), file=f.name)
        r = text_migrate_file(
            f, sorted_mapping,
            dry_run=dry_run,
            backup=backup,
            case_insensitive=case_insensitive,
            dense_threshold=dense_threshold,
            on_progress=on_progress,
        )
        results.append(r)

    ok      = sum(1 for r in results if r.success and r.changed)
    no_chg  = sum(1 for r in results if r.success and not r.changed)
    errors  = sum(1 for r in results if not r.success)
    total_c = sum(len(r.changes) for r in results)

    _emit(on_progress,
          "success" if not errors else "warning",
          f"Xong: {ok} migrate, {no_chg} không đổi, {errors} lỗi — {total_c} thay thế")

    return results


# ──────────────────────────────────────────────────────────────────────────────
# Summary formatter  (cùng interface với migrator.py)
# ──────────────────────────────────────────────────────────────────────────────

def text_format_summary(results: list[TextFileResult], dry_run: bool = False) -> str:
    lines = []
    for r in results:
        if not r.success:
            lines.append(f"❌ {r.path.name}: {r.error}")
        elif not r.changed:
            lines.append(f"⬜ {r.path.name}: không có text mẫu nào khớp")
        else:
            icon = "🔍" if dry_run else "✅"
            lines.append(f"{icon} {r.path.name}: {len(r.changes)} thay thế")
            seen: set[str] = set()
            for c in r.changes[:10]:
                key = f"{c.original}→{c.placeholder}"
                if key not in seen:
                    lines.append(f"     \"{c.original}\"  →  {c.placeholder}")
                    seen.add(key)
            if len(r.changes) > 10:
                lines.append(f"     ... và {len(r.changes) - 10} thay thế khác")
        if r.warnings:
            for w in r.warnings:
                lines.append(f"     ⚠ {w}")
    return "\n".join(lines) if lines else "(Không có file nào)"
